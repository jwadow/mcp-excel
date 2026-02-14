# Excel MCP Server
# Copyright (C) 2026 Jwadow
# Licensed under AGPL-3.0
# https://github.com/jwadow/mcp-excel

"""File loader with automatic format detection and caching."""

import os
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd

from .cache import FileCache
from .datetime_converter import DateTimeConverter
from .datetime_detector import DateTimeDetector


class FileLoader:
    """Loads Excel files with automatic format detection and caching."""

    def __init__(self, cache: Optional[FileCache] = None) -> None:
        """Initialize file loader.

        Args:
            cache: Optional FileCache instance. If None, creates default cache.
        """
        self._cache = cache or FileCache()
        self._datetime_detector = DateTimeDetector()
        self._datetime_converter = DateTimeConverter()

    def _detect_format(self, file_path: Path) -> str:
        """Detect Excel file format from extension.

        Args:
            file_path: Path to the file

        Returns:
            File format: 'xls' or 'xlsx'

        Raises:
            ValueError: If file format is not supported
        """
        suffix = file_path.suffix.lower()
        if suffix == ".xls":
            return "xls"
        elif suffix == ".xlsx":
            return "xlsx"
        else:
            raise ValueError(f"Unsupported file format: {suffix}. Only .xls and .xlsx are supported.")

    def _get_engine(self, file_format: str) -> str:
        """Get appropriate pandas engine for file format.

        Args:
            file_format: File format ('xls' or 'xlsx')

        Returns:
            Engine name for pandas
        """
        if file_format == "xls":
            return "xlrd"
        else:
            return "openpyxl"

    def load(
        self,
        file_path: str | Path,
        sheet_name: Optional[str | int] = 0,
        header_row: Optional[int] = None,
        use_cache: bool = True,
        convert_dates: bool = True,
    ) -> pd.DataFrame:
        """Load Excel file into DataFrame.

        Args:
            file_path: Absolute path to the Excel file
            sheet_name: Sheet name or index (default: 0 = first sheet)
            header_row: Row index to use as header (None = auto-detect)
            use_cache: Whether to use cache (default: True)
            convert_dates: Automatically detect and convert date columns (default: True)

        Returns:
            Loaded DataFrame

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If file format is unsupported
            Exception: If file cannot be read
        """
        path = Path(file_path)

        if not path.exists():
            raise FileNotFoundError(
                f"File not found: {file_path}\n"
                f"Please use an absolute path to the file."
            )
        
        # Try cache first
        if use_cache:
            sheet_key = str(sheet_name) if sheet_name is not None else "0"
            # Include header_row and convert_dates in cache key
            cache_key = f"{sheet_key}::header_{header_row}::dates_{convert_dates}"
            cached_df = self._cache.get(path, cache_key)
            if cached_df is not None:
                return cached_df

        # Detect format and engine
        file_format = self._detect_format(path)
        engine = self._get_engine(file_format)

        # Load file with specified header (None = raw data, int = specific row)
        try:
            df = pd.read_excel(
                path,
                sheet_name=sheet_name,
                engine=engine,
                header=header_row,
            )

            # Convert dates if requested
            if convert_dates:
                df = self._convert_datetime_columns(df, path, sheet_name, file_format)
            
            # Cache the result
            if use_cache:
                sheet_key = str(sheet_name) if sheet_name is not None else "0"
                cache_key = f"{sheet_key}::header_{header_row}::dates_{convert_dates}"
                self._cache.put(path, df, cache_key)

            return df

        except Exception as e:
            raise Exception(f"Failed to load file {file_path}: {str(e)}") from e

    def get_sheet_names(self, file_path: str | Path) -> list[str]:
        """Get list of sheet names in Excel file.

        Args:
            file_path: Absolute path to the Excel file

        Returns:
            List of sheet names

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If file format is unsupported
        """
        path = Path(file_path)

        if not path.exists():
            raise FileNotFoundError(
                f"File not found: {file_path}\n"
                f"Please use an absolute path to the file."
            )

        file_format = self._detect_format(path)
        engine = self._get_engine(file_format)

        try:
            excel_file = pd.ExcelFile(path, engine=engine)
            return excel_file.sheet_names
        except Exception as e:
            raise Exception(f"Failed to read sheet names from {file_path}: {str(e)}") from e

    def get_file_info(self, file_path: str | Path) -> dict[str, any]:
        """Get basic information about Excel file.

        Args:
            file_path: Absolute path to the Excel file

        Returns:
            Dictionary with file information

        Raises:
            FileNotFoundError: If file doesn't exist
        """
        path = Path(file_path)

        if not path.exists():
            raise FileNotFoundError(
                f"File not found: {file_path}\n"
                f"Please use an absolute path to the file."
            )

        file_format = self._detect_format(path)
        file_size = os.path.getsize(path)
        sheet_names = self.get_sheet_names(path)

        return {
            "format": file_format,
            "size_bytes": file_size,
            "size_mb": round(file_size / 1024 / 1024, 2),
            "sheet_count": len(sheet_names),
            "sheet_names": sheet_names,
        }

    def invalidate_cache(self, file_path: str | Path) -> None:
        """Invalidate cache for specific file.

        Args:
            file_path: Path to the file
        """
        self._cache.invalidate(Path(file_path))

    def clear_cache(self) -> None:
        """Clear entire cache."""
        self._cache.clear()

    def _convert_datetime_columns(
        self,
        df: pd.DataFrame,
        file_path: Path,
        sheet_name: Optional[str | int],
        file_format: str
    ) -> pd.DataFrame:
        """Convert datetime columns in DataFrame.
        
        Args:
            df: DataFrame to process
            file_path: Path to Excel file
            sheet_name: Sheet name or index
            file_format: File format ('xls' or 'xlsx')
        
        Returns:
            DataFrame with converted datetime columns
        """
        # Extract cell formats (only for .xlsx)
        cell_formats = None
        if file_format == "xlsx":
            try:
                cell_formats = self._extract_cell_formats_xlsx(file_path, sheet_name)
            except Exception:
                # If format extraction fails, fall back to heuristics
                pass
        
        # Detect datetime columns
        datetime_cols = self._datetime_detector.detect_datetime_columns(df, cell_formats)
        
        # Convert detected columns
        for col_name, date_info in datetime_cols.items():
            try:
                if date_info.source == "pandas_dtype":
                    # Pandas already parsed dates, just ensure consistent datetime64 type
                    # This handles both datetime64 columns and object columns with datetime objects
                    df[col_name] = pd.to_datetime(df[col_name], errors='coerce')
                elif date_info.source in ["cell_format", "heuristic"]:
                    # Need to convert from Excel numbers to datetime
                    # Detect epoch (Windows vs Mac)
                    epoch = self._datetime_converter.detect_epoch(df[col_name])
                    # Convert column
                    df[col_name] = self._datetime_converter.convert_column(df[col_name], epoch)
            except Exception:
                # If conversion fails, leave column as-is
                pass
        
        return df
    
    def _extract_cell_formats_xlsx(
        self,
        file_path: Path,
        sheet_name: Optional[str | int]
    ) -> Dict[str, List[str]]:
        """Extract cell formats from .xlsx file.
        
        Args:
            file_path: Path to .xlsx file
            sheet_name: Sheet name or index
        
        Returns:
            Dictionary mapping column names to format strings
        """
        try:
            import openpyxl
        except ImportError:
            # openpyxl not available, return empty dict
            return {}
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
            
            # Get worksheet
            if isinstance(sheet_name, int):
                ws = wb.worksheets[sheet_name]
            elif isinstance(sheet_name, str):
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            formats = {}
            
            # Read formats from first data row (row 2, assuming row 1 is header)
            # We sample multiple rows to get more reliable format detection
            for row_idx in range(2, min(12, ws.max_row + 1)):  # Sample up to 10 data rows
                for col_idx, cell in enumerate(ws[row_idx], start=0):
                    if cell.number_format and cell.number_format != 'General':
                        col_key = str(col_idx)
                        if col_key not in formats:
                            formats[col_key] = []
                        if cell.number_format not in formats[col_key]:
                            formats[col_key].append(cell.number_format)
            
            wb.close()
            return formats
            
        except Exception:
            return {}
    
    def get_cache_stats(self) -> dict[str, any]:
        """Get cache statistics.

        Returns:
            Dictionary with cache stats
        """
        return self._cache.get_stats()
