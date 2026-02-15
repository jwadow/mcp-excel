# Excel MCP Server
# Copyright (C) 2026 Jwadow
# Licensed under AGPL-3.0
# https://github.com/jwadow/mcp-excel

"""Test Excel fixtures generator.

This script creates a set of synthetic Excel files for testing.
Run ONCE to generate fixtures, then commit them to git.

Usage:
    python tests/builders/generate_fixtures.py
"""

import sys
from datetime import datetime, timedelta
from pathlib import Path

# Add src to path for imports (if running directly)
sys.path.insert(0, str(Path(__file__).parent.parent.parent / "src"))

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

try:
    import xlwt
except ImportError:
    print("⚠️ Warning: xlwt not installed. Legacy .xls files will be skipped.")
    print("   To generate .xls files, run: pip install xlwt")
    xlwt = None


class ExcelFixtureBuilder:
    """Builder for creating test Excel files."""

    def __init__(self, fixtures_root: Path):
        self.fixtures_root = fixtures_root
        # Create subdirectories
        self.basic_dir = fixtures_root / "basic"
        self.messy_dir = fixtures_root / "messy"
        self.edge_cases_dir = fixtures_root / "edge_cases"
        self.legacy_dir = fixtures_root / "legacy"
        
        # Ensure all directories exist
        for dir_path in [self.basic_dir, self.messy_dir, self.edge_cases_dir, self.legacy_dir]:
            dir_path.mkdir(parents=True, exist_ok=True)

    def create_simple_xlsx(self) -> Path:
        """Creates simple table: 3 columns, 10 rows, header in row 1.
        
        Uses Cyrillic data to test encoding handling.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"

        # Headers (Cyrillic)
        headers = ["Имя", "Возраст", "Город"]
        ws.append(headers)

        # Data (Cyrillic names and cities)
        data = [
            ["Алексей", 25, "Москва"],
            ["Мария", 30, "Лондон"],
            ["Дмитрий", 35, "Нью-Йорк"],
            ["Елена", 28, "Париж"],
            ["Иван", 32, "Токио"],
            ["Ольга", 27, "Берлин"],
            ["Сергей", 29, "Сидней"],
            ["Анна", 31, "Торонто"],
            ["Павел", 26, "Мадрид"],
            ["Наталья", 33, "Рим"],
        ]
        for row in data:
            ws.append(row)

        output_path = self.basic_dir / "simple.xlsx"
        wb.save(output_path)
        return output_path

    def create_with_dates_xlsx(self) -> Path:
        """Creates table with datetime columns.
        
        Tests datetime detection and conversion.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales"

        # Headers (Cyrillic)
        headers = ["Номер заказа", "Клиент", "Сумма", "Дата заказа", "Дата доставки"]
        ws.append(headers)

        # Data with dates
        base_date = datetime(2024, 1, 1, 10, 30)  # With time component
        clients = ["Ромашка", "Лютик", "Василёк", "Одуванчик", "Подснежник"]
        
        for i in range(1, 16):
            order_date = base_date + timedelta(days=i * 2, hours=i % 24)
            delivery_date = order_date + timedelta(days=3, hours=2)
            ws.append([
                f"ЗАК-{1000 + i}",
                clients[i % len(clients)],
                1000 + i * 100,
                order_date,
                delivery_date
            ])

        # Format date columns
        for row in range(2, ws.max_row + 1):
            ws.cell(row, 4).number_format = "DD/MM/YYYY HH:MM"
            ws.cell(row, 5).number_format = "DD/MM/YYYY HH:MM"

        output_path = self.basic_dir / "with_dates.xlsx"
        wb.save(output_path)
        return output_path

    def create_numeric_types_xlsx(self) -> Path:
        """Creates table with different numeric types (int, float).
        
        Tests numeric type detection and formatting.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Numbers"

        # Headers (Cyrillic)
        headers = ["Код товара", "Количество", "Цена", "Скидка", "Итого"]
        ws.append(headers)

        # Data: int, int, float, float, float
        for i in range(1, 21):
            product_id = 50089400 + i  # Large integers (test formatting)
            quantity = i * 10
            price = 99.99 + i * 5.5
            discount = 0.05 + (i % 5) * 0.02
            total = quantity * price * (1 - discount)
            ws.append([product_id, quantity, price, discount, total])

        output_path = self.basic_dir / "numeric_types.xlsx"
        wb.save(output_path)
        return output_path

    def create_messy_headers_xlsx(self) -> Path:
        """Creates table with headers starting from row 3 (real world scenario).
        
        Tests header detection algorithm.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"

        # Junk in first rows (like in enterprise files)
        ws.append(["ООО 'Рога и Копыта'"])
        ws.append(["Отчёт за январь 2024"])
        ws.append([])  # Empty row

        # Headers in row 4 (index 3)
        headers = ["Клиент", "Сумма", "Дата", "Статус"]
        ws.append(headers)

        # Data
        base_date = datetime(2024, 1, 1)
        clients = ["Ромашка", "Лютик", "Василёк", "Одуванчик", "Подснежник"]
        statuses = ["Выполнен", "В работе", "Отменён"]

        for i in range(20):
            ws.append([
                clients[i % len(clients)],
                1000 + i * 150,
                base_date + timedelta(days=i),
                statuses[i % len(statuses)]
            ])

        output_path = self.messy_dir / "messy_headers.xlsx"
        wb.save(output_path)
        return output_path

    def create_with_nulls_xlsx(self) -> Path:
        """Creates table with null/empty values.
        
        Tests null handling and find_nulls operation.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"

        # Headers (Cyrillic)
        headers = ["ID", "Имя", "Email", "Телефон", "Примечания"]
        ws.append(headers)

        # Data with nulls
        data = [
            [1, "Алексей", "alex@example.com", "123-456", "VIP клиент"],
            [2, "Мария", None, "234-567", None],  # No email and notes
            [3, "Дмитрий", "dmitry@example.com", None, "Новый клиент"],  # No phone
            [4, None, "unknown@example.com", "345-678", None],  # No name
            [5, "Елена", "elena@example.com", "456-789", "Постоянный"],
            [6, "Иван", None, None, None],  # Only ID and name
            [7, "Ольга", "olga@example.com", "567-890", None],
            [8, "Сергей", "sergey@example.com", None, "VIP клиент"],
            [9, None, None, "678-901", "Анонимный"],  # No name and email
            [10, "Анна", "anna@example.com", "789-012", "Постоянный"],
        ]
        for row in data:
            ws.append(row)

        output_path = self.edge_cases_dir / "with_nulls.xlsx"
        wb.save(output_path)
        return output_path

    def create_with_duplicates_xlsx(self) -> Path:
        """Creates table with duplicates for testing find_duplicates.
        
        Tests duplicate detection with various scenarios.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Orders"

        # Headers (Cyrillic)
        headers = ["Клиент", "Товар", "Количество", "Дата"]
        ws.append(headers)

        # Data with intentional duplicates
        base_date = datetime(2024, 1, 1)
        data = [
            ["Алексей", "Ноутбук", 1, base_date],
            ["Мария", "Мышь", 2, base_date + timedelta(days=1)],
            ["Алексей", "Ноутбук", 1, base_date],  # Duplicate of row 1
            ["Дмитрий", "Клавиатура", 1, base_date + timedelta(days=2)],
            ["Мария", "Мышь", 2, base_date + timedelta(days=1)],  # Duplicate of row 2
            ["Елена", "Монитор", 1, base_date + timedelta(days=3)],
            ["Алексей", "Ноутбук", 1, base_date],  # Another duplicate of row 1
            ["Иван", "Наушники", 1, base_date + timedelta(days=4)],
        ]
        for row in data:
            ws.append(row)

        output_path = self.edge_cases_dir / "with_duplicates.xlsx"
        wb.save(output_path)
        return output_path

    def create_wide_table_xlsx(self) -> Path:
        """Creates wide table (50 columns) for edge case testing.
        
        Tests handling of tables with many columns.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Wide"

        # 50 columns
        num_cols = 50
        headers = [f"Колонка_{i+1}" for i in range(num_cols)]
        ws.append(headers)

        # 10 rows of data
        for row_idx in range(10):
            row_data = [f"Значение_{row_idx}_{col_idx}" for col_idx in range(num_cols)]
            ws.append(row_data)

        output_path = self.edge_cases_dir / "wide_table.xlsx"
        wb.save(output_path)
        return output_path

    def create_single_column_xlsx(self) -> Path:
        """Creates table with single column (edge case).
        
        Tests handling of minimal table structure.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Single"

        # Single column (Cyrillic)
        ws.append(["Значение"])
        for i in range(1, 11):
            ws.append([f"Элемент {i}"])

        output_path = self.edge_cases_dir / "single_column.xlsx"
        wb.save(output_path)
        return output_path

    def create_mixed_languages_xlsx(self) -> Path:
        """Creates table with mixed Cyrillic, Latin, and special characters.
        
        Tests unicode handling and encoding edge cases.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mixed"

        # Headers (mixed)
        headers = ["Name/Имя", "Age/Возраст", "City/Город", "Comment/Комментарий"]
        ws.append(headers)

        # Data with mixed languages and special chars
        data = [
            ["Алексей/Alex", 25, "Москва/Moscow", "Обычный клиент"],
            ["Мария/Maria", 30, "Санкт-Петербург", "VIP 🌟"],
            ["John/Джон", 35, "New York/Нью-Йорк", "Discount 10% / Скидка 10%"],
            ["Елена/Elena", 28, "Екатеринбург", "New client ✓ / Новый клиент ✓"],
            ["Иван/Ivan", 32, "Казань/Kazan", "Regular customer / Постоянный покупатель"],
            ["François/Франсуа", 29, "Paris/Париж", "Spécial caractères: é, è, ê, ë"],
            ["李明/Li Ming", 31, "北京/Beijing", "中文测试 / Chinese test"],
            ["José/Хосе", 27, "Madrid/Мадрид", "¡Hola! ¿Cómo estás?"],
        ]
        for row in data:
            ws.append(row)

        output_path = self.edge_cases_dir / "mixed_languages.xlsx"
        wb.save(output_path)
        return output_path

    def create_special_chars_xlsx(self) -> Path:
        """Creates table with special characters and edge case strings.
        
        Tests formula injection protection and special char handling.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Special"

        # Headers
        headers = ["ID", "Текст", "Спецсимволы"]
        ws.append(headers)

        # Data with special characters
        data = [
            [1, "=1+1", "Formula injection test"],
            [2, "+7 (999) 123-45-67", "Phone with plus"],
            [3, "-100", "Negative number as text"],
            [4, "@username", "At symbol"],
            [5, "Текст с \"кавычками\"", "Quotes test"],
            [6, "Строка\nс переносом", "Newline test"],
            [7, "Табуляция\tздесь", "Tab test"],
            [8, "100%", "Percent symbol"],
            [9, "Цена: $99.99", "Dollar sign"],
            [10, "Email: test@example.com", "At in email"],
        ]
        for row in data:
            ws.append(row)

        output_path = self.edge_cases_dir / "special_chars.xlsx"
        wb.save(output_path)
        return output_path

    def create_merged_cells_xlsx(self) -> Path:
        """Creates table with merged cells in headers (common in reports).
        
        Tests handling of merged cells and complex header structures.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"

        # Merged header cells (typical enterprise report)
        ws['A1'] = "Отчёт о продажах за 2024 год"
        ws.merge_cells('A1:E1')  # Title across 5 columns
        
        ws['A2'] = "Регион"
        ws.merge_cells('A2:A3')  # Vertical merge
        
        ws['B2'] = "Квартал 1"
        ws.merge_cells('B2:C2')  # Horizontal merge for Q1
        ws['B3'] = "Январь"
        ws['C3'] = "Февраль"
        
        ws['D2'] = "Квартал 2"
        ws.merge_cells('D2:E2')  # Horizontal merge for Q2
        ws['D3'] = "Март"
        ws['E3'] = "Апрель"

        # Data rows
        regions = ["Москва", "Санкт-Петербург", "Новосибирск", "Екатеринбург", "Казань"]
        for i, region in enumerate(regions, start=4):
            ws.cell(i, 1, region)
            ws.cell(i, 2, 1000 + i * 100)
            ws.cell(i, 3, 1200 + i * 120)
            ws.cell(i, 4, 1100 + i * 110)
            ws.cell(i, 5, 1300 + i * 130)

        output_path = self.messy_dir / "merged_cells.xlsx"
        wb.save(output_path)
        return output_path

    def create_multilevel_headers_xlsx(self) -> Path:
        """Creates table with 3-level headers (deep hierarchy).
        
        Tests complex multi-level header detection.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales"

        # Level 1: Company name
        ws['A1'] = "ООО 'Рога и Копыта' - Годовой отчёт"
        ws.merge_cells('A1:G1')

        # Level 2: Main categories
        ws['A2'] = "Информация"
        ws.merge_cells('A2:B2')
        ws['C2'] = "Продажи"
        ws.merge_cells('C2:E2')
        ws['F2'] = "Финансы"
        ws.merge_cells('F2:G2')

        # Level 3: Subcategories
        ws['A3'] = "ID"
        ws['B3'] = "Клиент"
        ws['C3'] = "Q1"
        ws['D3'] = "Q2"
        ws['E3'] = "Q3"
        ws['F3'] = "Доход"
        ws['G3'] = "Расход"

        # Data
        for i in range(10):
            ws.cell(i + 4, 1, f"ID-{1000 + i}")
            ws.cell(i + 4, 2, f"Клиент {chr(65 + i % 5)}")
            ws.cell(i + 4, 3, 1000 + i * 50)
            ws.cell(i + 4, 4, 1200 + i * 60)
            ws.cell(i + 4, 5, 1100 + i * 55)
            ws.cell(i + 4, 6, 3300 + i * 165)
            ws.cell(i + 4, 7, 2000 + i * 100)

        output_path = self.messy_dir / "multilevel_headers.xlsx"
        wb.save(output_path)
        return output_path

    def create_enterprise_chaos_xlsx(self) -> Path:
        """Creates ultra-complex enterprise report (worst case scenario).
        
        Combines: junk rows, merged cells, multi-level headers, empty rows, mixed data.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Отчёт"

        # Row 1-2: Company header (junk)
        ws['A1'] = "ООО 'Рога и Копыта'"
        ws.merge_cells('A1:F1')
        ws['A2'] = "ИНН: 1234567890, КПП: 123456789"
        ws.merge_cells('A2:F2')

        # Row 3: Empty
        
        # Row 4: Report title
        ws['A4'] = "Сводный отчёт по продажам и закупкам за январь-март 2024"
        ws.merge_cells('A4:F4')

        # Row 5: Empty

        # Row 6-7: Multi-level headers with merges
        ws['A6'] = "Контрагент"
        ws.merge_cells('A6:A7')
        
        ws['B6'] = "Продажи"
        ws.merge_cells('B6:D6')
        ws['B7'] = "Январь"
        ws['C7'] = "Февраль"
        ws['D7'] = "Март"
        
        ws['E6'] = "Закупки"
        ws.merge_cells('E6:F6')
        ws['E7'] = "Сумма"
        ws['F7'] = "Количество"

        # Row 8: Data starts
        clients = ["Ромашка", "Лютик", "Василёк", "Одуванчик", "Подснежник"]
        for i, client in enumerate(clients, start=8):
            ws.cell(i, 1, client)
            ws.cell(i, 2, 1000 + i * 100)
            ws.cell(i, 3, 1200 + i * 120)
            ws.cell(i, 4, 1100 + i * 110)
            ws.cell(i, 5, 5000 + i * 500)
            ws.cell(i, 6, 50 + i * 5)

        # Row 13: Empty
        
        # Row 14: Footer with merged cells
        ws['A14'] = "Итого:"
        ws['B14'] = "=SUM(B8:B12)"
        ws['C14'] = "=SUM(C8:C12)"
        ws['D14'] = "=SUM(D8:D12)"
        ws['E14'] = "=SUM(E8:E12)"
        ws['F14'] = "=SUM(F8:F12)"

        output_path = self.messy_dir / "enterprise_chaos.xlsx"
        wb.save(output_path)
        return output_path

    def create_with_formulas_xlsx(self) -> Path:
        """Creates table with Excel formulas in cells.
        
        Tests formula handling and calculation.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Calculations"

        # Headers
        headers = ["Товар", "Цена", "Количество", "Сумма", "НДС 20%", "Итого"]
        ws.append(headers)

        # Data with formulas
        products = ["Ноутбук", "Мышь", "Клавиатура", "Монитор", "Наушники"]
        prices = [50000, 1500, 3000, 20000, 5000]
        quantities = [2, 10, 5, 3, 8]

        for i, (product, price, qty) in enumerate(zip(products, prices, quantities), start=2):
            ws.cell(i, 1, product)
            ws.cell(i, 2, price)
            ws.cell(i, 3, qty)
            ws.cell(i, 4, f"=B{i}*C{i}")  # Formula: Price * Quantity
            ws.cell(i, 5, f"=D{i}*0.2")   # Formula: Sum * 20%
            ws.cell(i, 6, f"=D{i}+E{i}")  # Formula: Sum + VAT

        # Total row with formulas
        ws.cell(7, 1, "ИТОГО:")
        ws.cell(7, 4, "=SUM(D2:D6)")
        ws.cell(7, 5, "=SUM(E2:E6)")
        ws.cell(7, 6, "=SUM(F2:F6)")

        output_path = self.edge_cases_dir / "with_formulas.xlsx"
        wb.save(output_path)
        return output_path

    def create_complex_formatting_xlsx(self) -> Path:
        """Creates table with various number formats.
        
        Tests number format detection and handling.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Formats"

        # Headers
        headers = ["Описание", "Значение", "Формат"]
        ws.append(headers)

        # Data with different formats
        data = [
            ("Целое число", 12345, "General"),
            ("Дробное число", 123.45, "0.00"),
            ("Процент", 0.15, "0.00%"),
            ("Валюта", 1234.56, "#,##0.00 ₽"),
            ("Дата", datetime(2024, 3, 15), "DD/MM/YYYY"),
            ("Время", datetime(2024, 1, 1, 14, 30), "HH:MM:SS"),
            ("Дата и время", datetime(2024, 3, 15, 14, 30), "DD/MM/YYYY HH:MM"),
            ("Научная нотация", 1.23e10, "0.00E+00"),
            ("Дробь", 0.75, "# ?/?"),
            ("Телефон", "+7 (999) 123-45-67", "@"),
        ]

        for row_idx, (desc, value, fmt) in enumerate(data, start=2):
            ws.cell(row_idx, 1, desc)
            cell = ws.cell(row_idx, 2, value)
            ws.cell(row_idx, 3, fmt)
            
            # Apply format
            if fmt == "0.00%":
                cell.number_format = "0.00%"
            elif fmt == "#,##0.00 ₽":
                cell.number_format = "#,##0.00 ₽"
            elif fmt == "DD/MM/YYYY":
                cell.number_format = "DD/MM/YYYY"
            elif fmt == "HH:MM:SS":
                cell.number_format = "HH:MM:SS"
            elif fmt == "DD/MM/YYYY HH:MM":
                cell.number_format = "DD/MM/YYYY HH:MM"
            elif fmt == "0.00E+00":
                cell.number_format = "0.00E+00"
            elif fmt == "# ?/?":
                cell.number_format = "# ?/?"

        output_path = self.edge_cases_dir / "complex_formatting.xlsx"
        wb.save(output_path)
        return output_path

    def create_multi_sheet_xlsx(self) -> Path:
        """Creates file with multiple sheets for multi-sheet testing.
        
        Tests multi-sheet operations and cache separation.
        """
        wb = openpyxl.Workbook()
        
        # Sheet 1: Products
        ws1 = wb.active
        ws1.title = "Products"
        ws1.append(["Товар", "Цена", "Категория"])
        products = [
            ["Ноутбук", 50000, "Электроника"],
            ["Мышь", 1500, "Электроника"],
            ["Стол", 15000, "Мебель"],
            ["Стул", 5000, "Мебель"],
            ["Книга", 500, "Книги"],
        ]
        for row in products:
            ws1.append(row)
        
        # Sheet 2: Clients
        ws2 = wb.create_sheet("Clients")
        ws2.append(["Клиент", "Город", "Рейтинг"])
        clients = [
            ["Ромашка", "Москва", 5],
            ["Лютик", "Санкт-Петербург", 4],
            ["Василёк", "Казань", 5],
            ["Одуванчик", "Екатеринбург", 3],
        ]
        for row in clients:
            ws2.append(row)
        
        # Sheet 3: Orders
        ws3 = wb.create_sheet("Orders")
        ws3.append(["Номер", "Клиент", "Товар", "Количество"])
        orders = [
            ["ЗАК-001", "Ромашка", "Ноутбук", 2],
            ["ЗАК-002", "Лютик", "Мышь", 5],
            ["ЗАК-003", "Василёк", "Стол", 1],
        ]
        for row in orders:
            ws3.append(row)
        
        output_path = self.basic_dir / "multi_sheet.xlsx"
        wb.save(output_path)
        return output_path

    def create_simple_xls(self) -> Path:
        """Creates simple table in legacy .xls format.
        
        Tests xlrd engine and legacy format support.
        """
        if xlwt is None:
            print("  ⚠️ Skipping .xls generation (xlwt not installed)")
            return None

        wb = xlwt.Workbook()
        ws = wb.add_sheet("Data")

        # Headers (Cyrillic)
        headers = ["Имя", "Возраст", "Город"]
        for col_idx, header in enumerate(headers):
            ws.write(0, col_idx, header)

        # Data (Cyrillic)
        data = [
            ["Алексей", 25, "Москва"],
            ["Мария", 30, "Лондон"],
            ["Дмитрий", 35, "Нью-Йорк"],
            ["Елена", 28, "Париж"],
            ["Иван", 32, "Токио"],
        ]
        for row_idx, row in enumerate(data, start=1):
            for col_idx, value in enumerate(row):
                ws.write(row_idx, col_idx, value)

        output_path = self.legacy_dir / "simple_legacy.xls"
        wb.save(str(output_path))
        return output_path


def main():
    """Generates all test fixtures."""
    print("=" * 80)
    print("  Excel Test Fixtures Generator")
    print("=" * 80)
    print()

    # Define paths
    script_dir = Path(__file__).parent
    fixtures_dir = script_dir.parent / "fixtures"

    builder = ExcelFixtureBuilder(fixtures_dir)

    fixtures_created = []

    # Generate fixtures
    print("📊 Generating fixtures...\n")

    # Basic fixtures
    print("1️⃣ Basic fixtures:")
    fixtures_created.append(("simple.xlsx", builder.create_simple_xlsx()))
    print(f"  ✅ simple.xlsx - simple table (3 columns, 10 rows, Cyrillic data)")

    fixtures_created.append(("with_dates.xlsx", builder.create_with_dates_xlsx()))
    print(f"  ✅ with_dates.xlsx - table with datetime columns")

    fixtures_created.append(("numeric_types.xlsx", builder.create_numeric_types_xlsx()))
    print(f"  ✅ numeric_types.xlsx - different numeric types (int, float)")
    
    fixtures_created.append(("multi_sheet.xlsx", builder.create_multi_sheet_xlsx()))
    print(f"  ✅ multi_sheet.xlsx - file with 3 sheets (Products, Clients, Orders)")

    # Messy fixtures (real world)
    print("\n2️⃣ Messy fixtures (real world scenarios):")
    fixtures_created.append(("messy_headers.xlsx", builder.create_messy_headers_xlsx()))
    print(f"  ✅ messy_headers.xlsx - headers from row 4, junk above")

    fixtures_created.append(("merged_cells.xlsx", builder.create_merged_cells_xlsx()))
    print(f"  ✅ merged_cells.xlsx - merged cells in headers (enterprise reports)")

    fixtures_created.append(("multilevel_headers.xlsx", builder.create_multilevel_headers_xlsx()))
    print(f"  ✅ multilevel_headers.xlsx - 3-level header hierarchy")

    fixtures_created.append(("enterprise_chaos.xlsx", builder.create_enterprise_chaos_xlsx()))
    print(f"  ✅ enterprise_chaos.xlsx - worst case: junk + merged + multi-level + formulas")

    # Edge cases
    print("\n3️⃣ Edge cases:")
    fixtures_created.append(("with_nulls.xlsx", builder.create_with_nulls_xlsx()))
    print(f"  ✅ with_nulls.xlsx - table with null/empty values")

    fixtures_created.append(("with_duplicates.xlsx", builder.create_with_duplicates_xlsx()))
    print(f"  ✅ with_duplicates.xlsx - table with duplicate rows")

    fixtures_created.append(("wide_table.xlsx", builder.create_wide_table_xlsx()))
    print(f"  ✅ wide_table.xlsx - wide table (50 columns)")

    fixtures_created.append(("single_column.xlsx", builder.create_single_column_xlsx()))
    print(f"  ✅ single_column.xlsx - single column (edge case)")

    fixtures_created.append(("mixed_languages.xlsx", builder.create_mixed_languages_xlsx()))
    print(f"  ✅ mixed_languages.xlsx - Cyrillic, Latin, Chinese, special chars")

    fixtures_created.append(("special_chars.xlsx", builder.create_special_chars_xlsx()))
    print(f"  ✅ special_chars.xlsx - formula injection tests, special symbols")

    fixtures_created.append(("with_formulas.xlsx", builder.create_with_formulas_xlsx()))
    print(f"  ✅ with_formulas.xlsx - Excel formulas in cells")

    fixtures_created.append(("complex_formatting.xlsx", builder.create_complex_formatting_xlsx()))
    print(f"  ✅ complex_formatting.xlsx - various number formats (%, currency, dates)")

    # Legacy format
    print("\n4️⃣ Legacy format (.xls):")
    legacy_path = builder.create_simple_xls()
    if legacy_path:
        fixtures_created.append(("simple_legacy.xls", legacy_path))
        print(f"  ✅ simple_legacy.xls - legacy format for xlrd testing")
    else:
        print(f"  ⚠️ simple_legacy.xls - skipped (xlwt not installed)")

    # Summary
    print("\n" + "=" * 80)
    print(f"✅ Created {len(fixtures_created)} fixtures in {fixtures_dir}")
    print("=" * 80)
    print("\n📋 Next steps:")
    print("  1. Check files in tests/fixtures/")
    print("  2. Open several files in Excel to verify")
    print("  3. Commit fixtures: git add tests/fixtures/")
    print("  4. Tests will use these static files")
    print()


if __name__ == "__main__":
    main()
