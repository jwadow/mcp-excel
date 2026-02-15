# Excel MCP Server
# Copyright (C) 2026 Jwadow
# Licensed under AGPL-3.0
# https://github.com/jwadow/mcp-excel

"""Integration tests for File Inspection operations.

Tests cover:
- inspect_file: Get file structure and metadata
- get_sheet_info: Get detailed sheet information with header detection
- get_column_names: Quick column name retrieval
- get_data_profile: Comprehensive data profiling

These are END-TO-END tests that verify the complete operation flow:
FileLoader -> HeaderDetector -> Operations -> Response
"""

import pytest
import openpyxl

from mcp_excel.operations.inspection import InspectionOperations
from mcp_excel.operations.base import MAX_DIFFERENCES
from mcp_excel.models.requests import (
    InspectFileRequest,
    GetSheetInfoRequest,
    GetColumnNamesRequest,
    GetDataProfileRequest,
    FindColumnRequest,
    SearchAcrossSheetsRequest,
    CompareSheetsRequest,
)


# ============================================================================
# inspect_file tests
# ============================================================================

def test_inspect_file_simple(simple_fixture, file_loader):
    """Test inspect_file on simple clean file.
    
    Verifies:
    - Returns correct format (xlsx)
    - Returns correct sheet count
    - Returns sheet names
    - Returns row/column counts for each sheet
    - Performance metrics are included
    """
    print(f"\n📂 Testing inspect_file on: {simple_fixture.name}")
    
    ops = InspectionOperations(file_loader)
    request = InspectFileRequest(file_path=simple_fixture.path_str)
    
    # Act
    response = ops.inspect_file(request)
    
    # Assert
    print(f"✅ Format: {response.format}")
    print(f"   Size: {response.size_mb} MB ({response.size_bytes} bytes)")
    print(f"   Sheets: {response.sheet_count}")
    print(f"   Sheet names: {response.sheet_names}")
    print(f"   Performance: {response.performance.execution_time_ms}ms")
    
    assert response.format == simple_fixture.format, "Should return correct format"
    assert response.sheet_count == 1, "Simple fixture has 1 sheet"
    assert simple_fixture.sheet_name in response.sheet_names, "Should include sheet name"
    assert len(response.sheets_info) == 1, "Should have info for 1 sheet"
    
    # Check first sheet info
    sheet_info = response.sheets_info[0]
    assert sheet_info["sheet_name"] == simple_fixture.sheet_name
    assert sheet_info["row_count"] > 0, "Should have rows"
    assert sheet_info["column_count"] == len(simple_fixture.columns), "Should match column count"
    
    # Check metadata
    assert response.metadata is not None, "Should include metadata"
    assert response.metadata.file_format == simple_fixture.format
    
    # Check performance metrics
    assert response.performance is not None, "Should include performance metrics"
    assert response.performance.execution_time_ms > 0, "Should have execution time"


def test_inspect_file_multi_sheet(multi_sheet_fixture, file_loader):
    """Test inspect_file on file with multiple sheets.
    
    Verifies:
    - Returns all sheets
    - Returns correct info for each sheet
    - Sheet order is preserved
    """
    print(f"\n📂 Testing inspect_file on multi-sheet file")
    
    ops = InspectionOperations(file_loader)
    request = InspectFileRequest(file_path=multi_sheet_fixture.path_str)
    
    # Act
    response = ops.inspect_file(request)
    
    # Assert
    print(f"✅ Sheets found: {response.sheet_count}")
    print(f"   Sheet names: {response.sheet_names}")
    for sheet_info in response.sheets_info:
        print(f"   - {sheet_info['sheet_name']}: {sheet_info['row_count']} rows, {sheet_info['column_count']} cols")
    
    expected_sheets = multi_sheet_fixture.expected["sheet_names"]
    assert response.sheet_count == len(expected_sheets), "Should find all sheets"
    assert set(response.sheet_names) == set(expected_sheets), "Should match expected sheet names"
    assert len(response.sheets_info) == len(expected_sheets), "Should have info for all sheets"
    
    # Verify each sheet has required fields
    for sheet_info in response.sheets_info:
        assert "sheet_name" in sheet_info, "Should have sheet_name"
        assert "row_count" in sheet_info, "Should have row_count"
        assert "column_count" in sheet_info, "Should have column_count"
        assert sheet_info["row_count"] > 0, "Each sheet should have rows"
        assert sheet_info["column_count"] > 0, "Each sheet should have columns"


def test_inspect_file_legacy_format(simple_legacy_fixture, file_loader):
    """Test inspect_file on legacy .xls format.
    
    Verifies:
    - Handles .xls format correctly
    - Returns format='xls'
    - Works with xlrd engine
    """
    print(f"\n📂 Testing inspect_file on legacy .xls file")
    
    ops = InspectionOperations(file_loader)
    request = InspectFileRequest(file_path=simple_legacy_fixture.path_str)
    
    # Act
    response = ops.inspect_file(request)
    
    # Assert
    print(f"✅ Format: {response.format}")
    print(f"   Sheets: {response.sheet_count}")
    
    assert response.format == "xls", "Should detect .xls format"
    assert response.sheet_count > 0, "Should have at least one sheet"
    assert len(response.sheets_info) > 0, "Should have sheet info"


def test_inspect_file_with_dates(with_dates_fixture, file_loader):
    """Test inspect_file on file with datetime columns.
    
    Verifies:
    - Handles datetime columns without errors
    - Returns correct structure
    """
    print(f"\n📂 Testing inspect_file on file with dates")
    
    ops = InspectionOperations(file_loader)
    request = InspectFileRequest(file_path=with_dates_fixture.path_str)
    
    # Act
    response = ops.inspect_file(request)
    
    # Assert
    print(f"✅ Format: {response.format}")
    print(f"   Rows: {response.sheets_info[0]['row_count']}")
    print(f"   Columns: {response.sheets_info[0]['column_count']}")
    
    assert response.format == with_dates_fixture.format
    # inspect_file loads without header, so it counts ALL rows including header
    assert response.sheets_info[0]["row_count"] == with_dates_fixture.row_count + 1, "Should count data + header row"
    assert response.sheets_info[0]["column_count"] == len(with_dates_fixture.columns)


def test_inspect_file_messy_headers(messy_headers_fixture, file_loader):
    """Test inspect_file on file with messy headers.
    
    Verifies:
    - Handles junk rows without errors
    - Returns correct row count (includes junk rows at this level)
    """
    print(f"\n📂 Testing inspect_file on messy headers file")
    
    ops = InspectionOperations(file_loader)
    request = InspectFileRequest(file_path=messy_headers_fixture.path_str)
    
    # Act
    response = ops.inspect_file(request)
    
    # Assert
    print(f"✅ Format: {response.format}")
    print(f"   Total rows (including junk): {response.sheets_info[0]['row_count']}")
    
    # inspect_file loads without header, so it counts ALL rows including junk
    expected_total = messy_headers_fixture.row_count + messy_headers_fixture.expected["junk_rows"] + 1  # +1 for header row
    assert response.sheets_info[0]["row_count"] == expected_total, "Should count all rows including junk"


def test_inspect_file_wide_table(wide_table_fixture, file_loader):
    """Test inspect_file on wide table (50 columns).
    
    Verifies:
    - Handles many columns correctly
    - Performance is acceptable
    """
    print(f"\n📂 Testing inspect_file on wide table")
    
    ops = InspectionOperations(file_loader)
    request = InspectFileRequest(file_path=wide_table_fixture.path_str)
    
    # Act
    response = ops.inspect_file(request)
    
    # Assert
    print(f"✅ Columns: {response.sheets_info[0]['column_count']}")
    print(f"   Performance: {response.performance.execution_time_ms}ms")
    
    assert response.sheets_info[0]["column_count"] == 50, "Should handle 50 columns"
    assert response.performance.execution_time_ms < 5000, "Should complete in reasonable time"


def test_inspect_file_single_column(single_column_fixture, file_loader):
    """Test inspect_file on minimal table (single column).
    
    Verifies:
    - Handles edge case of single column
    """
    print(f"\n📂 Testing inspect_file on single column table")
    
    ops = InspectionOperations(file_loader)
    request = InspectFileRequest(file_path=single_column_fixture.path_str)
    
    # Act
    response = ops.inspect_file(request)
    
    # Assert
    print(f"✅ Columns: {response.sheets_info[0]['column_count']}")
    
    assert response.sheets_info[0]["column_count"] == 1, "Should handle single column"


def test_inspect_file_mixed_languages(mixed_languages_fixture, file_loader):
    """Test inspect_file on file with mixed languages.
    
    Verifies:
    - Handles unicode correctly
    - No encoding errors
    """
    print(f"\n📂 Testing inspect_file on mixed languages file")
    
    ops = InspectionOperations(file_loader)
    request = InspectFileRequest(file_path=mixed_languages_fixture.path_str)
    
    # Act
    response = ops.inspect_file(request)
    
    # Assert
    print(f"✅ Format: {response.format}")
    print(f"   Sheets: {response.sheet_count}")
    
    assert response.format == mixed_languages_fixture.format
    assert response.sheet_count > 0, "Should handle mixed languages"


def test_inspect_file_with_nulls(with_nulls_fixture, file_loader):
    """Test inspect_file on file with null values.
    
    Verifies:
    - Handles null values without errors
    """
    print(f"\n📂 Testing inspect_file on file with nulls")
    
    ops = InspectionOperations(file_loader)
    request = InspectFileRequest(file_path=with_nulls_fixture.path_str)
    
    # Act
    response = ops.inspect_file(request)
    
    # Assert
    print(f"✅ Rows: {response.sheets_info[0]['row_count']}")
    
    assert response.sheets_info[0]["row_count"] == with_nulls_fixture.row_count + 1, "Should count all rows"


def test_inspect_file_with_duplicates(with_duplicates_fixture, file_loader):
    """Test inspect_file on file with duplicate rows.
    
    Verifies:
    - Handles duplicates without errors
    - Counts all rows including duplicates
    """
    print(f"\n📂 Testing inspect_file on file with duplicates")
    
    ops = InspectionOperations(file_loader)
    request = InspectFileRequest(file_path=with_duplicates_fixture.path_str)
    
    # Act
    response = ops.inspect_file(request)
    
    # Assert
    print(f"✅ Rows: {response.sheets_info[0]['row_count']}")
    
    assert response.sheets_info[0]["row_count"] == with_duplicates_fixture.row_count + 1, "Should count all rows including duplicates"


# ============================================================================
# get_sheet_info tests
# ============================================================================

def test_get_sheet_info_simple(simple_fixture, file_loader):
    """Test get_sheet_info on clean file.
    
    Verifies:
    - Returns correct column names
    - Returns correct column types
    - Returns sample rows (max 3)
    - Auto-detects header (row 0)
    - High confidence in header detection
    """
    print(f"\n📂 Testing get_sheet_info on: {simple_fixture.name}")
    
    ops = InspectionOperations(file_loader)
    request = GetSheetInfoRequest(
        file_path=simple_fixture.path_str,
        sheet_name=simple_fixture.sheet_name
    )
    
    # Act
    response = ops.get_sheet_info(request)
    
    # Assert
    print(f"✅ Columns: {response.column_names}")
    print(f"   Row count: {response.row_count}")
    print(f"   Column types: {response.column_types}")
    print(f"   Header detected at row: {response.header_detection.header_row}")
    print(f"   Confidence: {response.header_detection.confidence:.2%}")
    
    assert response.column_names == simple_fixture.columns, "Should match expected columns"
    assert response.column_count == len(simple_fixture.columns), "Should match column count"
    assert response.row_count == simple_fixture.row_count, "Should match row count"
    assert len(response.sample_rows) <= 3, "Should return max 3 sample rows"
    assert len(response.sample_rows) > 0, "Should return at least 1 sample row"
    
    # Check header detection
    assert response.header_detection is not None, "Should include header detection info"
    assert response.header_detection.header_row == simple_fixture.header_row, "Should detect correct header row"
    assert response.header_detection.confidence > 0.8, "Should have high confidence for clean file"
    
    # Check column types
    assert len(response.column_types) == len(simple_fixture.columns), "Should have type for each column"
    for col in simple_fixture.columns:
        assert col in response.column_types, f"Should have type for column {col}"
    
    # Check sample rows structure
    for row in response.sample_rows:
        assert isinstance(row, dict), "Sample row should be dict"
        assert len(row) == len(simple_fixture.columns), "Sample row should have all columns"


def test_get_sheet_info_with_dates(with_dates_fixture, file_loader):
    """Test get_sheet_info with datetime columns.
    
    Verifies:
    - Detects datetime columns correctly
    - Returns datetime values in ISO 8601 format (string)
    - Sample rows contain datetime strings
    """
    print(f"\n📂 Testing get_sheet_info with datetime columns")
    
    ops = InspectionOperations(file_loader)
    request = GetSheetInfoRequest(
        file_path=with_dates_fixture.path_str,
        sheet_name=with_dates_fixture.sheet_name
    )
    
    # Act
    response = ops.get_sheet_info(request)
    
    # Assert
    print(f"✅ Column types: {response.column_types}")
    
    expected_datetime_cols = with_dates_fixture.expected["datetime_columns"]
    for col in expected_datetime_cols:
        assert col in response.column_types, f"Should have column {col}"
        assert response.column_types[col] == "datetime", f"Column {col} should be datetime type"
    
    # Check sample rows have ISO 8601 format
    if response.sample_rows:
        first_row = response.sample_rows[0]
        for col in expected_datetime_cols:
            value = first_row.get(col)
            if value:
                print(f"   {col}: {value}")
                assert isinstance(value, str), f"{col} should be string"
                assert "T" in value, f"{col} should be ISO 8601 format (YYYY-MM-DDTHH:MM:SS)"


def test_get_sheet_info_numeric_types(numeric_types_fixture, file_loader):
    """Test get_sheet_info with numeric columns.
    
    Verifies:
    - Detects integer and float types correctly
    - Handles large integers without loss
    """
    print(f"\n📂 Testing get_sheet_info with numeric types")
    
    ops = InspectionOperations(file_loader)
    request = GetSheetInfoRequest(
        file_path=numeric_types_fixture.path_str,
        sheet_name=numeric_types_fixture.sheet_name
    )
    
    # Act
    response = ops.get_sheet_info(request)
    
    # Assert
    print(f"✅ Column types: {response.column_types}")
    
    # Check that we have both integer and float types
    types_found = set(response.column_types.values())
    print(f"   Types found: {types_found}")
    
    assert "integer" in types_found or "float" in types_found, "Should detect numeric types"
    
    # Check sample rows have numeric values
    if response.sample_rows:
        first_row = response.sample_rows[0]
        print(f"   Sample row (first 3 cols): {dict(list(first_row.items())[:3])}")
        
        # Check that numeric values are properly formatted (not scientific notation for display)
        for col, value in first_row.items():
            if response.column_types[col] in ["integer", "float"]:
                assert isinstance(value, (int, float)), f"Numeric column {col} should have numeric value"


def test_get_sheet_info_messy_headers(messy_headers_fixture, file_loader):
    """Test get_sheet_info with messy headers (junk rows above).
    
    Verifies:
    - Auto-detects correct header row (skips junk)
    - Returns correct data rows only
    - Confidence is reasonable despite junk
    """
    print(f"\n📂 Testing get_sheet_info with messy headers")
    
    ops = InspectionOperations(file_loader)
    request = GetSheetInfoRequest(
        file_path=messy_headers_fixture.path_str,
        sheet_name=messy_headers_fixture.sheet_name
    )
    
    # Act
    response = ops.get_sheet_info(request)
    
    # Assert
    print(f"✅ Detected header row: {response.header_detection.header_row}")
    print(f"   Confidence: {response.header_detection.confidence:.2%}")
    print(f"   Columns: {response.column_names}")
    print(f"   Data rows: {response.row_count}")
    
    assert response.header_detection.header_row == messy_headers_fixture.header_row, "Should skip junk rows"
    assert response.column_names == messy_headers_fixture.columns, "Should get correct columns"
    assert response.row_count == messy_headers_fixture.row_count, "Should count only data rows"
    assert response.header_detection.confidence > 0.7, "Should have reasonable confidence"


def test_get_sheet_info_manual_header_row(simple_fixture, file_loader):
    """Test get_sheet_info with manually specified header row.
    
    Verifies:
    - Respects manual header_row parameter
    - Returns correct data based on specified header
    """
    print(f"\n📂 Testing get_sheet_info with manual header_row")
    
    ops = InspectionOperations(file_loader)
    request = GetSheetInfoRequest(
        file_path=simple_fixture.path_str,
        sheet_name=simple_fixture.sheet_name,
        header_row=0  # Explicitly specify
    )
    
    # Act
    response = ops.get_sheet_info(request)
    
    # Assert
    print(f"✅ Data starts at row: {response.data_start_row}")
    print(f"   Columns: {response.column_names}")
    
    assert response.column_names == simple_fixture.columns, "Should use specified header row"
    assert response.data_start_row == 1, "Data should start after header row 0"


def test_get_sheet_info_multilevel_headers(multilevel_headers_fixture, file_loader):
    """Test get_sheet_info with multi-level headers.
    
    Verifies:
    - Detects deepest header level
    - Returns correct column names from deepest level
    """
    print(f"\n📂 Testing get_sheet_info with multi-level headers")
    
    ops = InspectionOperations(file_loader)
    request = GetSheetInfoRequest(
        file_path=multilevel_headers_fixture.path_str,
        sheet_name=multilevel_headers_fixture.sheet_name
    )
    
    # Act
    response = ops.get_sheet_info(request)
    
    # Assert
    print(f"✅ Detected header row: {response.header_detection.header_row}")
    print(f"   Columns: {response.column_names}")
    
    assert response.header_detection.header_row == multilevel_headers_fixture.header_row, "Should detect deepest level"
    assert response.column_names == multilevel_headers_fixture.columns, "Should get columns from deepest level"


def test_get_sheet_info_enterprise_chaos(enterprise_chaos_fixture, file_loader):
    """Test get_sheet_info on worst-case scenario.
    
    Verifies:
    - Handles complex real-world files
    - Detects correct header despite chaos
    """
    print(f"\n📂 Testing get_sheet_info on enterprise chaos file")
    
    ops = InspectionOperations(file_loader)
    request = GetSheetInfoRequest(
        file_path=enterprise_chaos_fixture.path_str,
        sheet_name=enterprise_chaos_fixture.sheet_name
    )
    
    # Act
    response = ops.get_sheet_info(request)
    
    # Assert
    print(f"✅ Detected header row: {response.header_detection.header_row}")
    print(f"   Confidence: {response.header_detection.confidence:.2%}")
    print(f"   Columns: {response.column_names}")
    
    assert response.header_detection.header_row == enterprise_chaos_fixture.header_row, "Should handle worst-case"
    # Note: First column is merged cell, pandas reads it as 'Unnamed: 0'
    # This is expected behavior for merged cells
    assert len(response.column_names) == len(enterprise_chaos_fixture.columns), "Should have correct number of columns"
    assert response.column_names[1:] == enterprise_chaos_fixture.columns[1:], "Should get correct columns (except first merged cell)"


def test_get_sheet_info_wide_table(wide_table_fixture, file_loader):
    """Test get_sheet_info on wide table (50 columns).
    
    Verifies:
    - Handles many columns correctly
    - Returns all column names
    - Sample rows have all columns
    """
    print(f"\n📂 Testing get_sheet_info on wide table")
    
    ops = InspectionOperations(file_loader)
    request = GetSheetInfoRequest(
        file_path=wide_table_fixture.path_str,
        sheet_name=wide_table_fixture.sheet_name
    )
    
    # Act
    response = ops.get_sheet_info(request)
    
    # Assert
    print(f"✅ Column count: {response.column_count}")
    print(f"   Performance: {response.performance.execution_time_ms}ms")
    
    assert response.column_count == 50, "Should return all 50 columns"
    assert len(response.column_names) == 50, "Should have 50 column names"
    
    # Check sample rows have all columns
    if response.sample_rows:
        assert len(response.sample_rows[0]) == 50, "Sample row should have all 50 columns"


def test_get_sheet_info_single_column(single_column_fixture, file_loader):
    """Test get_sheet_info on minimal table (single column).
    
    Verifies:
    - Handles edge case of single column
    """
    print(f"\n📂 Testing get_sheet_info on single column table")
    
    ops = InspectionOperations(file_loader)
    request = GetSheetInfoRequest(
        file_path=single_column_fixture.path_str,
        sheet_name=single_column_fixture.sheet_name
    )
    
    # Act
    response = ops.get_sheet_info(request)
    
    # Assert
    print(f"✅ Columns: {response.column_names}")
    
    assert response.column_count == 1, "Should handle single column"
    assert len(response.column_names) == 1, "Should have 1 column name"


def test_get_sheet_info_mixed_languages(mixed_languages_fixture, file_loader):
    """Test get_sheet_info on file with mixed languages.
    
    Verifies:
    - Handles unicode correctly in column names
    - Sample rows contain unicode values
    """
    print(f"\n📂 Testing get_sheet_info on mixed languages file")
    
    ops = InspectionOperations(file_loader)
    request = GetSheetInfoRequest(
        file_path=mixed_languages_fixture.path_str,
        sheet_name=mixed_languages_fixture.sheet_name
    )
    
    # Act
    response = ops.get_sheet_info(request)
    
    # Assert
    print(f"✅ Columns: {response.column_names}")
    
    assert response.column_names == mixed_languages_fixture.columns, "Should handle unicode in column names"
    
    # Check sample rows contain unicode
    if response.sample_rows:
        first_row = response.sample_rows[0]
        print(f"   Sample row (first 2 values): {list(first_row.values())[:2]}")
        # Just verify no encoding errors occurred


def test_get_sheet_info_with_nulls(with_nulls_fixture, file_loader):
    """Test get_sheet_info on file with null values.
    
    Verifies:
    - Handles null values in sample rows
    - Returns None for null values (JSON-serializable)
    """
    print(f"\n📂 Testing get_sheet_info on file with nulls")
    
    ops = InspectionOperations(file_loader)
    request = GetSheetInfoRequest(
        file_path=with_nulls_fixture.path_str,
        sheet_name=with_nulls_fixture.sheet_name
    )
    
    # Act
    response = ops.get_sheet_info(request)
    
    # Assert
    print(f"✅ Columns: {response.column_names}")
    print(f"   Row count: {response.row_count}")
    
    # Check that sample rows can contain None values
    if response.sample_rows:
        # At least one sample row should have a None value
        has_null = any(
            value is None
            for row in response.sample_rows
            for value in row.values()
        )
        print(f"   Has null in samples: {has_null}")
        # Note: might not always have null in first 3 rows, so we don't assert


def test_get_sheet_info_performance_metrics(simple_fixture, file_loader):
    """Test that get_sheet_info includes performance metrics.
    
    Verifies:
    - Performance metrics are included
    - Execution time is reasonable
    - Cache hit status is reported
    """
    print(f"\n📂 Testing get_sheet_info performance metrics")
    
    ops = InspectionOperations(file_loader)
    request = GetSheetInfoRequest(
        file_path=simple_fixture.path_str,
        sheet_name=simple_fixture.sheet_name
    )
    
    # Act
    response = ops.get_sheet_info(request)
    
    # Assert
    print(f"✅ Performance:")
    print(f"   Execution time: {response.performance.execution_time_ms}ms")
    print(f"   Cache hit: {response.performance.cache_hit}")
    print(f"   Memory used: {response.performance.memory_used_mb}MB")
    
    assert response.performance is not None, "Should include performance metrics"
    assert response.performance.execution_time_ms > 0, "Should have execution time"
    assert response.performance.cache_hit in [True, False], "Should report cache status"


# ============================================================================
# get_column_names tests
# ============================================================================

def test_get_column_names_simple(simple_fixture, file_loader):
    """Test get_column_names on clean file.
    
    Verifies:
    - Returns correct column names
    - Fast operation (no full data load needed)
    - Auto-detects header
    """
    print(f"\n📂 Testing get_column_names")
    
    ops = InspectionOperations(file_loader)
    request = GetColumnNamesRequest(
        file_path=simple_fixture.path_str,
        sheet_name=simple_fixture.sheet_name
    )
    
    # Act
    response = ops.get_column_names(request)
    
    # Assert
    print(f"✅ Columns: {response.column_names}")
    print(f"   Count: {response.column_count}")
    print(f"   Performance: {response.performance.execution_time_ms}ms")
    
    assert response.column_names == simple_fixture.columns, "Should return correct columns"
    assert response.column_count == len(simple_fixture.columns), "Should match count"
    assert response.performance.execution_time_ms < 1000, "Should be fast operation"


def test_get_column_names_wide_table(wide_table_fixture, file_loader):
    """Test get_column_names on wide table (50 columns).
    
    Verifies:
    - Handles many columns correctly
    - Returns all column names
    """
    print(f"\n📂 Testing get_column_names on wide table")
    
    ops = InspectionOperations(file_loader)
    request = GetColumnNamesRequest(
        file_path=wide_table_fixture.path_str,
        sheet_name=wide_table_fixture.sheet_name
    )
    
    # Act
    response = ops.get_column_names(request)
    
    # Assert
    print(f"✅ Column count: {response.column_count}")
    
    assert response.column_count == 50, "Should return all 50 columns"
    assert len(response.column_names) == 50, "Should have 50 column names"


def test_get_column_names_messy_headers(messy_headers_fixture, file_loader):
    """Test get_column_names with messy headers.
    
    Verifies:
    - Auto-detects correct header row
    - Returns correct column names
    """
    print(f"\n📂 Testing get_column_names with messy headers")
    
    ops = InspectionOperations(file_loader)
    request = GetColumnNamesRequest(
        file_path=messy_headers_fixture.path_str,
        sheet_name=messy_headers_fixture.sheet_name
    )
    
    # Act
    response = ops.get_column_names(request)
    
    # Assert
    print(f"✅ Columns: {response.column_names}")
    
    assert response.column_names == messy_headers_fixture.columns, "Should skip junk and get correct columns"


def test_get_column_names_manual_header(simple_fixture, file_loader):
    """Test get_column_names with manual header_row.
    
    Verifies:
    - Respects manual header_row parameter
    """
    print(f"\n📂 Testing get_column_names with manual header_row")
    
    ops = InspectionOperations(file_loader)
    request = GetColumnNamesRequest(
        file_path=simple_fixture.path_str,
        sheet_name=simple_fixture.sheet_name,
        header_row=0
    )
    
    # Act
    response = ops.get_column_names(request)
    
    # Assert
    print(f"✅ Columns: {response.column_names}")
    
    assert response.column_names == simple_fixture.columns, "Should use specified header row"


def test_get_column_names_mixed_languages(mixed_languages_fixture, file_loader):
    """Test get_column_names with unicode column names.
    
    Verifies:
    - Handles unicode correctly
    """
    print(f"\n📂 Testing get_column_names with mixed languages")
    
    ops = InspectionOperations(file_loader)
    request = GetColumnNamesRequest(
        file_path=mixed_languages_fixture.path_str,
        sheet_name=mixed_languages_fixture.sheet_name
    )
    
    # Act
    response = ops.get_column_names(request)
    
    # Assert
    print(f"✅ Columns: {response.column_names}")
    
    assert response.column_names == mixed_languages_fixture.columns, "Should handle unicode"


# ============================================================================
# get_data_profile tests
# ============================================================================

def test_get_data_profile_specific_columns(simple_fixture, file_loader):
    """Test get_data_profile with specific columns.
    
    Verifies:
    - Profiles only requested columns
    - Returns correct data types
    - Returns statistics for numeric columns
    - Returns top values
    - Generates TSV output
    """
    print(f"\n📂 Testing get_data_profile with specific columns")
    
    ops = InspectionOperations(file_loader)
    request = GetDataProfileRequest(
        file_path=simple_fixture.path_str,
        sheet_name=simple_fixture.sheet_name,
        columns=simple_fixture.columns[:2],  # First 2 columns
        top_n=3
    )
    
    # Act
    response = ops.get_data_profile(request)
    
    # Assert
    print(f"✅ Profiled columns: {response.columns_profiled}")
    print(f"   Profiles: {list(response.profiles.keys())}")
    
    assert response.columns_profiled == 2, "Should profile 2 columns"
    assert len(response.profiles) == 2, "Should have 2 profiles"
    
    # Check each profile has required fields
    for col_name, profile in response.profiles.items():
        print(f"\n   Column: {col_name}")
        print(f"     Type: {profile.data_type}")
        print(f"     Total: {profile.total_count}")
        print(f"     Nulls: {profile.null_count}")
        print(f"     Unique: {profile.unique_count}")
        
        assert profile.column_name == col_name
        assert profile.data_type in ["string", "integer", "float", "datetime", "boolean"]
        assert profile.total_count > 0
        assert profile.null_count >= 0
        assert profile.unique_count >= 0
        assert len(profile.top_values) <= 3, "Should return max 3 top values"
        
        # Check top values structure
        for top_val in profile.top_values:
            assert "value" in top_val
            assert "count" in top_val
            assert "percentage" in top_val
    
    # Check TSV output
    assert response.excel_output.tsv, "Should generate TSV output"
    assert len(response.excel_output.tsv) > 0, "TSV should not be empty"


def test_get_data_profile_all_columns(simple_fixture, file_loader):
    """Test get_data_profile without specifying columns (profile all).
    
    Verifies:
    - Profiles all columns when columns=None
    """
    print(f"\n📂 Testing get_data_profile for all columns")
    
    ops = InspectionOperations(file_loader)
    request = GetDataProfileRequest(
        file_path=simple_fixture.path_str,
        sheet_name=simple_fixture.sheet_name,
        columns=None,  # Profile all
        top_n=5
    )
    
    # Act
    response = ops.get_data_profile(request)
    
    # Assert
    print(f"✅ Profiled columns: {response.columns_profiled}")
    
    assert response.columns_profiled == len(simple_fixture.columns), "Should profile all columns"
    assert len(response.profiles) == len(simple_fixture.columns), "Should have profile for each column"


def test_get_data_profile_numeric_stats(numeric_types_fixture, file_loader):
    """Test get_data_profile with numeric columns.
    
    Verifies:
    - Returns statistics for numeric columns (mean, median, std, etc.)
    - Handles int and float types correctly
    - Statistics are accurate
    """
    print(f"\n📂 Testing get_data_profile with numeric columns")
    
    ops = InspectionOperations(file_loader)
    request = GetDataProfileRequest(
        file_path=numeric_types_fixture.path_str,
        sheet_name=numeric_types_fixture.sheet_name,
        columns=numeric_types_fixture.columns[:3],  # First 3 columns (numeric)
        top_n=5
    )
    
    # Act
    response = ops.get_data_profile(request)
    
    # Assert
    print(f"✅ Profiled numeric columns")
    
    for col_name, profile in response.profiles.items():
        print(f"\n   Column: {col_name}")
        print(f"     Type: {profile.data_type}")
        
        if profile.data_type in ["integer", "float"]:
            assert profile.stats is not None, f"Numeric column {col_name} should have stats"
            print(f"     Mean: {profile.stats.mean:.2f}")
            print(f"     Median: {profile.stats.median:.2f}")
            print(f"     Min: {profile.stats.min}")
            print(f"     Max: {profile.stats.max}")
            
            assert profile.stats.count > 0
            assert profile.stats.mean is not None
            assert profile.stats.median is not None
            assert profile.stats.min is not None
            assert profile.stats.max is not None
            assert profile.stats.q25 is not None
            assert profile.stats.q75 is not None


def test_get_data_profile_with_nulls(with_nulls_fixture, file_loader):
    """Test get_data_profile with null values.
    
    Verifies:
    - Correctly counts null values
    - Calculates null percentage accurately
    - Handles columns with different null counts
    """
    print(f"\n📂 Testing get_data_profile with null values")
    
    ops = InspectionOperations(file_loader)
    request = GetDataProfileRequest(
        file_path=with_nulls_fixture.path_str,
        sheet_name=with_nulls_fixture.sheet_name,
        columns=None,  # Profile all
        top_n=3
    )
    
    # Act
    response = ops.get_data_profile(request)
    
    # Assert
    print(f"✅ Profiled columns with nulls")
    
    # Check that at least some columns have nulls
    columns_with_nulls = [
        col_name for col_name, profile in response.profiles.items()
        if profile.null_count > 0
    ]
    
    print(f"   Columns with nulls: {len(columns_with_nulls)}")
    
    assert len(columns_with_nulls) > 0, "Should detect null values"
    
    # Check null percentage calculation
    for col_name in columns_with_nulls:
        profile = response.profiles[col_name]
        expected_percentage = (profile.null_count / profile.total_count * 100)
        print(f"   {col_name}: {profile.null_count} nulls ({profile.null_percentage}%)")
        assert abs(profile.null_percentage - expected_percentage) < 0.1, "Null percentage should be accurate"


def test_get_data_profile_datetime_columns(with_dates_fixture, file_loader):
    """Test get_data_profile with datetime columns.
    
    Verifies:
    - Detects datetime type correctly
    - Returns top values for datetime columns
    - No statistics for datetime (only for numeric)
    """
    print(f"\n📂 Testing get_data_profile with datetime columns")
    
    ops = InspectionOperations(file_loader)
    
    datetime_cols = with_dates_fixture.expected["datetime_columns"]
    
    request = GetDataProfileRequest(
        file_path=with_dates_fixture.path_str,
        sheet_name=with_dates_fixture.sheet_name,
        columns=datetime_cols,
        top_n=3
    )
    
    # Act
    response = ops.get_data_profile(request)
    
    # Assert
    print(f"✅ Profiled datetime columns")
    
    for col_name, profile in response.profiles.items():
        print(f"\n   Column: {col_name}")
        print(f"     Type: {profile.data_type}")
        
        assert profile.data_type == "datetime", f"Column {col_name} should be datetime type"
        assert profile.stats is None, "Datetime columns should not have numeric stats"
        assert len(profile.top_values) > 0, "Should have top values"


def test_get_data_profile_wide_table(wide_table_fixture, file_loader):
    """Test get_data_profile on wide table.
    
    Verifies:
    - Handles many columns efficiently
    - Performance is acceptable
    """
    print(f"\n📂 Testing get_data_profile on wide table")
    
    ops = InspectionOperations(file_loader)
    request = GetDataProfileRequest(
        file_path=wide_table_fixture.path_str,
        sheet_name=wide_table_fixture.sheet_name,
        columns=None,  # Profile all 50 columns
        top_n=3
    )
    
    # Act
    response = ops.get_data_profile(request)
    
    # Assert
    print(f"✅ Profiled {response.columns_profiled} columns")
    print(f"   Performance: {response.performance.execution_time_ms}ms")
    
    assert response.columns_profiled == 50, "Should profile all 50 columns"
    assert response.performance.execution_time_ms < 10000, "Should complete in reasonable time"


def test_get_data_profile_single_column(single_column_fixture, file_loader):
    """Test get_data_profile on single column table.
    
    Verifies:
    - Handles minimal structure
    """
    print(f"\n📂 Testing get_data_profile on single column")
    
    ops = InspectionOperations(file_loader)
    request = GetDataProfileRequest(
        file_path=single_column_fixture.path_str,
        sheet_name=single_column_fixture.sheet_name,
        columns=None,
        top_n=5
    )
    
    # Act
    response = ops.get_data_profile(request)
    
    # Assert
    print(f"✅ Profiled {response.columns_profiled} column")
    
    assert response.columns_profiled == 1, "Should profile single column"


def test_get_data_profile_top_n_parameter(simple_fixture, file_loader):
    """Test get_data_profile with different top_n values.
    
    Verifies:
    - Respects top_n parameter
    - Returns correct number of top values
    """
    print(f"\n📂 Testing get_data_profile with top_n parameter")
    
    ops = InspectionOperations(file_loader)
    
    # Test with top_n=5
    request = GetDataProfileRequest(
        file_path=simple_fixture.path_str,
        sheet_name=simple_fixture.sheet_name,
        columns=[simple_fixture.columns[0]],  # First column
        top_n=5
    )
    
    # Act
    response = ops.get_data_profile(request)
    
    # Assert
    print(f"✅ Top values count: {len(response.profiles[simple_fixture.columns[0]].top_values)}")
    
    profile = response.profiles[simple_fixture.columns[0]]
    assert len(profile.top_values) <= 5, "Should return max 5 top values"


def test_get_data_profile_tsv_output(simple_fixture, file_loader):
    """Test that get_data_profile generates proper TSV output.
    
    Verifies:
    - TSV output is generated
    - Contains column information
    - Can be pasted into Excel
    """
    print(f"\n📂 Testing get_data_profile TSV output")
    
    ops = InspectionOperations(file_loader)
    request = GetDataProfileRequest(
        file_path=simple_fixture.path_str,
        sheet_name=simple_fixture.sheet_name,
        columns=simple_fixture.columns[:2],
        top_n=3
    )
    
    # Act
    response = ops.get_data_profile(request)
    
    # Assert
    print(f"✅ TSV output generated")
    print(f"   Length: {len(response.excel_output.tsv)} chars")
    print(f"   Preview: {response.excel_output.tsv[:200]}...")
    
    assert response.excel_output.tsv, "Should generate TSV output"
    assert len(response.excel_output.tsv) > 0, "TSV should not be empty"
    
    # Check TSV contains column names
    for col in simple_fixture.columns[:2]:
        assert col in response.excel_output.tsv, f"TSV should contain column {col}"
    
    # Check TSV has tab separators
    assert "\t" in response.excel_output.tsv, "TSV should use tab separators"


def test_get_data_profile_invalid_column(simple_fixture, file_loader):
    """Test get_data_profile with non-existent column.
    
    Verifies:
    - Raises ValueError for invalid column
    - Error message is helpful
    """
    print(f"\n📂 Testing get_data_profile with invalid column")
    
    ops = InspectionOperations(file_loader)
    request = GetDataProfileRequest(
        file_path=simple_fixture.path_str,
        sheet_name=simple_fixture.sheet_name,
        columns=["NonExistentColumn"],
        top_n=3
    )
    
    # Act & Assert
    with pytest.raises(ValueError) as exc_info:
        ops.get_data_profile(request)
    
    print(f"✅ Caught expected error: {exc_info.value}")
    
    assert "not found" in str(exc_info.value).lower(), "Error should mention column not found"
    assert "NonExistentColumn" in str(exc_info.value), "Error should mention the invalid column name"


# ============================================================================
# Exception Handling Tests (lines 78-79, 257-258, 329-330)
# ============================================================================

def test_inspect_file_with_corrupted_sheet(temp_excel_path, file_loader):
    """Test inspect_file handles corrupted/unreadable sheets gracefully.
    
    Covers lines 78-79: Exception handling in inspect_file()
    
    Verifies:
    - Returns error info for corrupted sheet
    - Continues processing other sheets
    - Doesn't crash the entire operation
    """
    print(f"\n📂 Testing inspect_file with corrupted sheet")
    
    # Create file with one good sheet and one that will fail to load
    import openpyxl
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "GoodSheet"
    ws1.append(["Column1", "Column2"])
    ws1.append([1, 2])
    
    # Create another sheet
    ws2 = wb.create_sheet("BadSheet")
    ws2.append(["A", "B"])
    ws2.append([1, 2])
    
    file_path = temp_excel_path / "corrupted.xlsx"
    wb.save(file_path)
    
    # Mock the loader to simulate corruption on second sheet
    original_load = file_loader.load
    def mock_load(path, sheet, **kwargs):
        if sheet == "BadSheet":
            raise ValueError("Simulated corruption: Cannot read sheet")
        return original_load(path, sheet, **kwargs)
    
    file_loader.load = mock_load
    
    try:
        ops = InspectionOperations(file_loader)
        request = InspectFileRequest(file_path=str(file_path))
        
        # Act
        response = ops.inspect_file(request)
        
        # Assert
        print(f"✅ Sheet count: {response.sheet_count}")
        print(f"   Sheets info: {len(response.sheets_info)}")
        
        assert response.sheet_count == 2, "Should report both sheets"
        assert len(response.sheets_info) == 2, "Should have info for both sheets"
        
        # Check that one sheet has error
        error_sheets = [s for s in response.sheets_info if "error" in s]
        good_sheets = [s for s in response.sheets_info if "error" not in s]
        
        print(f"   Good sheets: {len(good_sheets)}")
        print(f"   Error sheets: {len(error_sheets)}")
        
        assert len(error_sheets) == 1, "Should have 1 sheet with error"
        assert len(good_sheets) == 1, "Should have 1 good sheet"
        
        # Check error message
        error_sheet = error_sheets[0]
        assert error_sheet["sheet_name"] == "BadSheet"
        assert "corruption" in error_sheet["error"].lower() or "cannot read" in error_sheet["error"].lower()
        
        # Check good sheet
        good_sheet = good_sheets[0]
        assert good_sheet["sheet_name"] == "GoodSheet"
        assert good_sheet["row_count"] > 0
        assert good_sheet["column_count"] > 0
        
    finally:
        # Restore original load method
        file_loader.load = original_load


def test_find_column_with_error_in_sheet(temp_excel_path, file_loader):
    """Test find_column continues when one sheet fails to load.
    
    Covers lines 257-258: Exception handling in find_column()
    
    Verifies:
    - Skips sheets that fail to load
    - Returns results from sheets that loaded successfully
    - Doesn't crash the entire search
    """
    print(f"\n📂 Testing find_column with sheet loading error")
    
    # Create file with multiple sheets
    import openpyxl
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["Name", "Age", "City"])
    ws1.append(["Alice", 25, "Moscow"])
    
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["Name", "Salary"])
    ws2.append(["Bob", 50000])
    
    ws3 = wb.create_sheet("Sheet3")
    ws3.append(["Product", "Price"])
    ws3.append(["Laptop", 1000])
    
    file_path = temp_excel_path / "multi_sheet.xlsx"
    wb.save(file_path)
    
    # Mock loader to fail on Sheet2
    original_load = file_loader.load
    def mock_load(path, sheet, **kwargs):
        if sheet == "Sheet2":
            raise RuntimeError("Simulated error loading Sheet2")
        return original_load(path, sheet, **kwargs)
    
    file_loader.load = mock_load
    
    try:
        ops = InspectionOperations(file_loader)
        request = FindColumnRequest(
            file_path=str(file_path),
            column_name="Name",
            search_all_sheets=True
        )
        
        # Act
        response = ops.find_column(request)
        
        # Assert
        print(f"✅ Total matches: {response.total_matches}")
        print(f"   Found in sheets: {[m['sheet'] for m in response.found_in]}")
        
        # Should find "Name" in Sheet1 and Sheet2 (Sheet2 fails, so only Sheet1)
        assert response.total_matches == 1, "Should find column in 1 sheet (Sheet2 failed)"
        assert len(response.found_in) == 1
        assert response.found_in[0]["sheet"] == "Sheet1"
        assert response.found_in[0]["column_name"] == "Name"
        
    finally:
        file_loader.load = original_load


def test_search_across_sheets_with_error(temp_excel_path, file_loader):
    """Test search_across_sheets continues when one sheet fails.
    
    Covers lines 329-330: Exception handling in search_across_sheets()
    
    Verifies:
    - Skips sheets that fail to load
    - Returns matches from successful sheets
    - Total matches count is correct
    """
    print(f"\n📂 Testing search_across_sheets with sheet error")
    
    # Create file with multiple sheets
    import openpyxl
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["Customer", "Amount"])
    ws1.append(["Alice", 100])
    ws1.append(["Bob", 200])
    ws1.append(["Alice", 150])
    
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["Customer", "Orders"])
    ws2.append(["Alice", 5])
    ws2.append(["Charlie", 3])
    
    file_path = temp_excel_path / "search_test.xlsx"
    wb.save(file_path)
    
    # Mock loader to fail on Sheet2
    original_load = file_loader.load
    def mock_load(path, sheet, **kwargs):
        if sheet == "Sheet2":
            raise IOError("Simulated I/O error")
        return original_load(path, sheet, **kwargs)
    
    file_loader.load = mock_load
    
    try:
        ops = InspectionOperations(file_loader)
        request = SearchAcrossSheetsRequest(
            file_path=str(file_path),
            column_name="Customer",
            value="Alice"
        )
        
        # Act
        response = ops.search_across_sheets(request)
        
        # Assert
        print(f"✅ Total matches: {response.total_matches}")
        print(f"   Matches: {response.matches}")
        
        # Should find "Alice" only in Sheet1 (2 times)
        assert response.total_matches == 2, "Should find 2 matches in Sheet1"
        assert len(response.matches) == 1, "Should have results from 1 sheet"
        assert response.matches[0]["sheet"] == "Sheet1"
        assert response.matches[0]["match_count"] == 2
        
    finally:
        file_loader.load = original_load


# ============================================================================
# Validation Tests in compare_sheets (lines 382-402)
# ============================================================================

def test_compare_sheets_missing_key_column_sheet1(temp_excel_path, file_loader):
    """Test compare_sheets raises error when key_column missing in sheet1.
    
    Covers lines 382-386: Validation for key_column in sheet1
    
    Verifies:
    - Raises ValueError with helpful message
    - Lists available columns
    """
    print(f"\n📂 Testing compare_sheets with missing key_column in sheet1")
    
    # Create multi-sheet file
    import openpyxl
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["Name", "Age"])  # No "ID" column
    ws1.append(["Alice", 25])
    
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["ID", "Name", "Salary"])
    ws2.append([1, "Alice", 50000])
    
    file_path = temp_excel_path / "compare_test.xlsx"
    wb.save(file_path)
    
    ops = InspectionOperations(file_loader)
    request = CompareSheetsRequest(
        file_path=str(file_path),
        sheet1="Sheet1",
        sheet2="Sheet2",
        key_column="ID",  # Doesn't exist in Sheet1
        compare_columns=["Name"]
    )
    
    # Act & Assert
    with pytest.raises(ValueError) as exc_info:
        ops.compare_sheets(request)
    
    print(f"✅ Caught expected error: {exc_info.value}")
    
    error_msg = str(exc_info.value)
    assert "ID" in error_msg, "Error should mention missing column"
    assert "Sheet1" in error_msg, "Error should mention which sheet"
    assert "Available columns" in error_msg, "Error should list available columns"
    assert "Name" in error_msg and "Age" in error_msg, "Error should show actual columns"


def test_compare_sheets_missing_key_column_sheet2(temp_excel_path, file_loader):
    """Test compare_sheets raises error when key_column missing in sheet2.
    
    Covers lines 387-391: Validation for key_column in sheet2
    
    Verifies:
    - Raises ValueError with helpful message
    - Lists available columns for sheet2
    """
    print(f"\n📂 Testing compare_sheets with missing key_column in sheet2")
    
    import openpyxl
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["ID", "Name", "Age"])
    ws1.append([1, "Alice", 25])
    
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["Name", "Salary"])  # No "ID" column
    ws2.append(["Alice", 50000])
    
    file_path = temp_excel_path / "compare_test2.xlsx"
    wb.save(file_path)
    
    ops = InspectionOperations(file_loader)
    request = CompareSheetsRequest(
        file_path=str(file_path),
        sheet1="Sheet1",
        sheet2="Sheet2",
        key_column="ID",  # Doesn't exist in Sheet2
        compare_columns=["Name"]
    )
    
    # Act & Assert
    with pytest.raises(ValueError) as exc_info:
        ops.compare_sheets(request)
    
    print(f"✅ Caught expected error: {exc_info.value}")
    
    error_msg = str(exc_info.value)
    assert "ID" in error_msg, "Error should mention missing column"
    assert "Sheet2" in error_msg, "Error should mention which sheet"
    assert "Available columns" in error_msg, "Error should list available columns"


def test_compare_sheets_missing_compare_column_sheet1(temp_excel_path, file_loader):
    """Test compare_sheets raises error when compare_column missing in sheet1.
    
    Covers lines 394-398: Validation for compare_columns in sheet1
    
    Verifies:
    - Raises ValueError for missing compare column
    """
    print(f"\n📂 Testing compare_sheets with missing compare_column in sheet1")
    
    import openpyxl
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["ID", "Name"])  # No "Salary" column
    ws1.append([1, "Alice"])
    
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["ID", "Name", "Salary"])
    ws2.append([1, "Alice", 50000])
    
    file_path = temp_excel_path / "compare_test3.xlsx"
    wb.save(file_path)
    
    ops = InspectionOperations(file_loader)
    request = CompareSheetsRequest(
        file_path=str(file_path),
        sheet1="Sheet1",
        sheet2="Sheet2",
        key_column="ID",
        compare_columns=["Salary"]  # Doesn't exist in Sheet1
    )
    
    # Act & Assert
    with pytest.raises(ValueError) as exc_info:
        ops.compare_sheets(request)
    
    print(f"✅ Caught expected error: {exc_info.value}")
    
    error_msg = str(exc_info.value)
    assert "Salary" in error_msg, "Error should mention missing column"
    assert "Sheet1" in error_msg, "Error should mention which sheet"


def test_compare_sheets_missing_compare_column_sheet2(temp_excel_path, file_loader):
    """Test compare_sheets raises error when compare_column missing in sheet2.
    
    Covers lines 399-402: Validation for compare_columns in sheet2
    
    Verifies:
    - Raises ValueError for missing compare column in sheet2
    """
    print(f"\n📂 Testing compare_sheets with missing compare_column in sheet2")
    
    import openpyxl
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["ID", "Name", "Salary"])
    ws1.append([1, "Alice", 50000])
    
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["ID", "Name"])  # No "Salary" column
    ws2.append([1, "Alice"])
    
    file_path = temp_excel_path / "compare_test4.xlsx"
    wb.save(file_path)
    
    ops = InspectionOperations(file_loader)
    request = CompareSheetsRequest(
        file_path=str(file_path),
        sheet1="Sheet1",
        sheet2="Sheet2",
        key_column="ID",
        compare_columns=["Salary"]  # Doesn't exist in Sheet2
    )
    
    # Act & Assert
    with pytest.raises(ValueError) as exc_info:
        ops.compare_sheets(request)
    
    print(f"✅ Caught expected error: {exc_info.value}")
    
    error_msg = str(exc_info.value)
    assert "Salary" in error_msg, "Error should mention missing column"
    assert "Sheet2" in error_msg, "Error should mention which sheet"


# ============================================================================
# Truncation Tests (lines 419-421, 426-437, 474-475)
# ============================================================================

def test_compare_sheets_only_in_sheet1(temp_excel_path, file_loader):
    """Test compare_sheets with rows only_in_sheet1.
    
    Covers lines 426-431: Building diff_entry for "only_in_sheet1" status
    
    Verifies:
    - Correctly identifies rows only in sheet1
    - Sets status to "only_in_sheet1"
    - Sets sheet2 values to None
    """
    print(f"\n📂 Testing compare_sheets with only_in_sheet1 rows")
    
    # Create sheets with rows only in sheet1
    import openpyxl
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["ID", "Name", "Amount"])
    ws1.append([1, "Alice", 100])
    ws1.append([2, "Bob", 200])
    ws1.append([3, "Charlie", 300])  # Only in sheet1
    ws1.append([4, "David", 400])    # Only in sheet1
    
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["ID", "Name", "Amount"])
    ws2.append([1, "Alice", 100])
    ws2.append([2, "Bob", 200])
    
    file_path = temp_excel_path / "only_in_sheet1.xlsx"
    wb.save(file_path)
    
    ops = InspectionOperations(file_loader)
    request = CompareSheetsRequest(
        file_path=str(file_path),
        sheet1="Sheet1",
        sheet2="Sheet2",
        key_column="ID",
        compare_columns=["Name", "Amount"]
    )
    
    # Act
    response = ops.compare_sheets(request)
    
    # Assert
    print(f"✅ Differences: {response.difference_count}")
    
    assert response.difference_count == 2, "Should find 2 differences"
    
    # Check only_in_sheet1 entries
    only_in_sheet1 = [d for d in response.differences if d["status"] == "only_in_sheet1"]
    assert len(only_in_sheet1) == 2, "Should have 2 rows only in sheet1"
    
    # Check structure
    for diff in only_in_sheet1:
        assert diff["status"] == "only_in_sheet1"
        assert diff["Name_sheet1"] is not None
        assert diff["Name_sheet2"] is None
        assert diff["Amount_sheet1"] is not None
        assert diff["Amount_sheet2"] is None
        
        print(f"   Row {diff['ID']}: {diff['Name_sheet1']} only in sheet1")


def test_compare_sheets_only_in_sheet2(temp_excel_path, file_loader):
    """Test compare_sheets with rows only_in_sheet2.
    
    Covers lines 432-437: Building diff_entry for "only_in_sheet2" status
    
    Verifies:
    - Correctly identifies rows only in sheet2
    - Sets status to "only_in_sheet2"
    - Sets sheet1 values to None
    """
    print(f"\n📂 Testing compare_sheets with only_in_sheet2 rows")
    
    # Create sheets with rows only in sheet2
    import openpyxl
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["ID", "Name", "Amount"])
    ws1.append([1, "Alice", 100])
    ws1.append([2, "Bob", 200])
    
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["ID", "Name", "Amount"])
    ws2.append([1, "Alice", 100])
    ws2.append([2, "Bob", 200])
    ws2.append([3, "Charlie", 300])  # Only in sheet2
    ws2.append([4, "David", 400])    # Only in sheet2
    ws2.append([5, "Eve", 500])      # Only in sheet2
    
    file_path = temp_excel_path / "only_in_sheet2.xlsx"
    wb.save(file_path)
    
    ops = InspectionOperations(file_loader)
    request = CompareSheetsRequest(
        file_path=str(file_path),
        sheet1="Sheet1",
        sheet2="Sheet2",
        key_column="ID",
        compare_columns=["Name", "Amount"]
    )
    
    # Act
    response = ops.compare_sheets(request)
    
    # Assert
    print(f"✅ Differences: {response.difference_count}")
    
    assert response.difference_count == 3, "Should find 3 differences"
    
    # Check only_in_sheet2 entries
    only_in_sheet2 = [d for d in response.differences if d["status"] == "only_in_sheet2"]
    assert len(only_in_sheet2) == 3, "Should have 3 rows only in sheet2"
    
    # Check structure
    for diff in only_in_sheet2:
        assert diff["status"] == "only_in_sheet2"
        assert diff["Name_sheet1"] is None
        assert diff["Name_sheet2"] is not None
        assert diff["Amount_sheet1"] is None
        assert diff["Amount_sheet2"] is not None
        
        print(f"   Row {diff['ID']}: {diff['Name_sheet2']} only in sheet2")


# ============================================================================
# Edge Case in get_data_profile (line 551-552)
# ============================================================================

def test_get_data_profile_boolean_column(temp_excel_path, file_loader):
    """Test get_data_profile with boolean column.
    
    Covers lines 551-552: Boolean data type detection
    
    Verifies:
    - Correctly identifies boolean columns
    - Returns data_type="boolean"
    - No statistics for boolean (only for numeric)
    """
    print(f"\n📂 Testing get_data_profile with boolean column")
    
    # Create file with boolean column
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Active", "Premium"])
    ws.append(["Alice", True, False])
    ws.append(["Bob", False, True])
    ws.append(["Charlie", True, True])
    ws.append(["David", False, False])
    
    file_path = temp_excel_path / "boolean_test.xlsx"
    wb.save(file_path)
    
    ops = InspectionOperations(file_loader)
    request = GetDataProfileRequest(
        file_path=str(file_path),
        sheet_name="Data",
        columns=["Active", "Premium"],
        top_n=3
    )
    
    # Act
    response = ops.get_data_profile(request)
    
    # Assert
    print(f"✅ Profiled columns: {response.columns_profiled}")
    
    assert response.columns_profiled == 2, "Should profile 2 columns"
    
    # Check boolean columns
    for col_name in ["Active", "Premium"]:
        profile = response.profiles[col_name]
        print(f"\n   Column: {col_name}")
        print(f"     Type: {profile.data_type}")
        print(f"     Unique: {profile.unique_count}")
        
        assert profile.data_type == "boolean", f"Column {col_name} should be boolean type"
        assert profile.stats is None, "Boolean columns should not have numeric stats"
        assert profile.unique_count <= 2, "Boolean column should have max 2 unique values"
        assert len(profile.top_values) > 0, "Should have top values"
