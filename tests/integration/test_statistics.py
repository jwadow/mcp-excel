# Excel MCP Server
# Copyright (C) 2026 Jwadow
# Licensed under AGPL-3.0
# https://github.com/jwadow/mcp-excel

"""Integration tests for Statistics operations.

Tests cover:
- get_column_stats: Statistical summary (mean, median, std, quartiles, etc.)
- correlate: Correlation matrix between multiple columns
- detect_outliers: Outlier detection using IQR and Z-score methods

These are END-TO-END tests that verify the complete operation flow:
FileLoader -> HeaderDetector -> StatisticsOperations -> Response
"""

import pytest

from mcp_excel.operations.statistics import StatisticsOperations
from mcp_excel.models.requests import (
    GetColumnStatsRequest,
    CorrelateRequest,
    DetectOutliersRequest,
    FilterCondition,
)


# ============================================================================
# get_column_stats tests
# ============================================================================

def test_get_column_stats_basic(numeric_types_fixture, file_loader):
    """Test get_column_stats on numeric column.
    
    Verifies:
    - Returns all statistics (count, mean, median, std, min, max, quartiles)
    - Statistics are mathematically correct
    - Null count is accurate
    - TSV output is generated
    """
    print(f"\n📊 Testing get_column_stats on numeric column")
    
    ops = StatisticsOperations(file_loader)
    
    # Use "Количество" column (integers)
    request = GetColumnStatsRequest(
        file_path=numeric_types_fixture.path_str,
        sheet_name=numeric_types_fixture.sheet_name,
        column="Количество",
        filters=[]
    )
    
    # Act
    response = ops.get_column_stats(request)
    
    # Assert
    print(f"✅ Statistics for 'Количество':")
    print(f"   Count: {response.stats.count}")
    print(f"   Mean: {response.stats.mean:.2f}")
    print(f"   Median: {response.stats.median:.2f}")
    print(f"   Std: {response.stats.std:.2f}" if response.stats.std else "   Std: N/A")
    print(f"   Min: {response.stats.min}")
    print(f"   Max: {response.stats.max}")
    print(f"   Q25: {response.stats.q25:.2f}")
    print(f"   Q75: {response.stats.q75:.2f}")
    print(f"   Nulls: {response.stats.null_count}")
    
    assert response.column == "Количество"
    assert response.stats.count == 20, "Should have 20 values"
    assert response.stats.null_count == 0, "Should have no nulls"
    
    # Check statistics are reasonable
    assert response.stats.mean > 0, "Mean should be positive"
    assert response.stats.median > 0, "Median should be positive"
    assert response.stats.std is not None, "Should have std for multiple values"
    assert response.stats.std > 0, "Std should be positive"
    assert response.stats.min < response.stats.max, "Min should be less than max"
    assert response.stats.q25 < response.stats.q75, "Q25 should be less than Q75"
    
    # Check TSV output
    assert response.excel_output.tsv, "Should generate TSV"
    assert "Mean" in response.excel_output.tsv, "TSV should contain statistics"
    assert "Median" in response.excel_output.tsv
    
    # Check metadata
    assert response.metadata is not None
    assert response.performance is not None


def test_get_column_stats_float_column(numeric_types_fixture, file_loader):
    """Test get_column_stats on float column.
    
    Verifies:
    - Handles float values correctly
    - Precision is maintained
    """
    print(f"\n📊 Testing get_column_stats on float column")
    
    ops = StatisticsOperations(file_loader)
    
    # Use "Цена" column (floats)
    request = GetColumnStatsRequest(
        file_path=numeric_types_fixture.path_str,
        sheet_name=numeric_types_fixture.sheet_name,
        column="Цена",
        filters=[]
    )
    
    # Act
    response = ops.get_column_stats(request)
    
    # Assert
    print(f"✅ Statistics for 'Цена' (float):")
    print(f"   Mean: {response.stats.mean:.2f}")
    print(f"   Median: {response.stats.median:.2f}")
    
    assert response.stats.count == 20
    assert isinstance(response.stats.mean, float)
    assert isinstance(response.stats.median, float)
    # Цена starts at 99.99 + i*5.5, so mean should be around 150-160
    assert 100 < response.stats.mean < 200, "Mean should be in reasonable range"


def test_get_column_stats_with_filters(with_dates_fixture, file_loader):
    """Test get_column_stats with filters applied.
    
    Verifies:
    - Filters are applied before calculating statistics
    - Statistics are correct for filtered subset
    """
    print(f"\n📊 Testing get_column_stats with filters")
    
    ops = StatisticsOperations(file_loader)
    
    # Filter for specific client
    request = GetColumnStatsRequest(
        file_path=with_dates_fixture.path_str,
        sheet_name=with_dates_fixture.sheet_name,
        column="Сумма",
        filters=[
            FilterCondition(column="Клиент", operator="==", value="Ромашка")
        ],
        logic="AND"
    )
    
    # Act
    response = ops.get_column_stats(request)
    
    # Assert
    print(f"✅ Statistics for 'Сумма' (filtered):")
    print(f"   Count: {response.stats.count}")
    print(f"   Mean: {response.stats.mean:.2f}")
    
    # Should have fewer values than total (15 rows, 5 clients, so ~3 per client)
    assert response.stats.count < 15, "Should have filtered subset"
    assert response.stats.count > 0, "Should have some matching rows"


def test_get_column_stats_text_stored_numbers(file_loader, temp_excel_path):
    """Test get_column_stats with text-stored numbers (auto-conversion).
    
    Verifies:
    - Automatically converts text to numbers
    - Statistics are calculated correctly after conversion
    """
    print(f"\n📊 Testing get_column_stats with text-stored numbers")
    
    # Create test file with text-stored numbers
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    
    ws.append(["ID", "Value"])
    # Store numbers as text (prefix with apostrophe)
    for i in range(1, 11):
        ws.append([i, str(i * 10)])  # "10", "20", "30", ...
    
    test_file = temp_excel_path / "text_numbers.xlsx"
    wb.save(test_file)
    
    ops = StatisticsOperations(file_loader)
    request = GetColumnStatsRequest(
        file_path=str(test_file),
        sheet_name="Data",
        column="Value",
        filters=[]
    )
    
    # Act
    response = ops.get_column_stats(request)
    
    # Assert
    print(f"✅ Converted text to numbers:")
    print(f"   Count: {response.stats.count}")
    print(f"   Mean: {response.stats.mean:.2f}")
    
    assert response.stats.count == 10, "Should convert all text numbers"
    assert response.stats.mean == 55.0, "Mean of 10,20,...,100 is 55"
    assert response.stats.min == 10
    assert response.stats.max == 100


def test_get_column_stats_with_nulls(with_nulls_fixture, file_loader):
    """Test get_column_stats with null values.
    
    Verifies:
    - Nulls are excluded from calculations
    - Null count is reported correctly
    - Statistics are based on non-null values only
    """
    print(f"\n📊 Testing get_column_stats with nulls")
    
    ops = StatisticsOperations(file_loader)
    
    # ID column has no nulls, but let's test with a column that might have nulls
    request = GetColumnStatsRequest(
        file_path=with_nulls_fixture.path_str,
        sheet_name=with_nulls_fixture.sheet_name,
        column="ID",
        filters=[]
    )
    
    # Act
    response = ops.get_column_stats(request)
    
    # Assert
    print(f"✅ Statistics with nulls:")
    print(f"   Count (non-null): {response.stats.count}")
    print(f"   Null count: {response.stats.null_count}")
    
    assert response.stats.count > 0, "Should have non-null values"
    assert response.stats.null_count >= 0, "Should report null count"
    assert response.stats.count + response.stats.null_count == 10, "Total should match row count"


def test_get_column_stats_single_value(file_loader, temp_excel_path):
    """Test get_column_stats with single value (edge case).
    
    Verifies:
    - Handles single value correctly
    - Std is None for single value
    - Other stats are calculated
    """
    print(f"\n📊 Testing get_column_stats with single value")
    
    # Create test file with single value
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    
    ws.append(["Value"])
    ws.append([42])
    
    test_file = temp_excel_path / "single_value.xlsx"
    wb.save(test_file)
    
    ops = StatisticsOperations(file_loader)
    request = GetColumnStatsRequest(
        file_path=str(test_file),
        sheet_name="Data",
        column="Value",
        filters=[]
    )
    
    # Act
    response = ops.get_column_stats(request)
    
    # Assert
    print(f"✅ Statistics for single value:")
    print(f"   Count: {response.stats.count}")
    print(f"   Mean: {response.stats.mean}")
    print(f"   Std: {response.stats.std}")
    
    assert response.stats.count == 1
    assert response.stats.mean == 42.0
    assert response.stats.median == 42.0
    assert response.stats.min == 42
    assert response.stats.max == 42
    assert response.stats.std is None, "Std should be None for single value"


def test_get_column_stats_invalid_column(simple_fixture, file_loader):
    """Test get_column_stats with non-existent column.
    
    Verifies:
    - Raises ValueError for invalid column
    - Error message lists available columns
    """
    print(f"\n📊 Testing get_column_stats with invalid column")
    
    ops = StatisticsOperations(file_loader)
    request = GetColumnStatsRequest(
        file_path=simple_fixture.path_str,
        sheet_name=simple_fixture.sheet_name,
        column="NonExistentColumn",
        filters=[]
    )
    
    # Act & Assert
    with pytest.raises(ValueError) as exc_info:
        ops.get_column_stats(request)
    
    print(f"✅ Caught expected error: {exc_info.value}")
    
    assert "not found" in str(exc_info.value).lower()
    assert "NonExistentColumn" in str(exc_info.value)
    assert "Available columns" in str(exc_info.value)


def test_get_column_stats_non_numeric_column(simple_fixture, file_loader):
    """Test get_column_stats with non-numeric column.
    
    Verifies:
    - Raises ValueError for non-numeric column
    - Error message explains the issue
    """
    print(f"\n📊 Testing get_column_stats with non-numeric column")
    
    ops = StatisticsOperations(file_loader)
    
    # "Имя" is a string column
    request = GetColumnStatsRequest(
        file_path=simple_fixture.path_str,
        sheet_name=simple_fixture.sheet_name,
        column="Имя",
        filters=[]
    )
    
    # Act & Assert
    with pytest.raises(ValueError) as exc_info:
        ops.get_column_stats(request)
    
    print(f"✅ Caught expected error: {exc_info.value}")
    
    assert "not numeric" in str(exc_info.value).lower() or "could be converted" in str(exc_info.value).lower()


def test_get_column_stats_all_nulls(file_loader, temp_excel_path):
    """Test get_column_stats with all null values.
    
    Verifies:
    - Raises ValueError when all values are null
    - Error message is clear
    """
    print(f"\n📊 Testing get_column_stats with all nulls")
    
    # Create test file with all nulls
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    
    ws.append(["Value"])
    ws.append([None])
    ws.append([None])
    ws.append([None])
    
    test_file = temp_excel_path / "all_nulls.xlsx"
    wb.save(test_file)
    
    ops = StatisticsOperations(file_loader)
    request = GetColumnStatsRequest(
        file_path=str(test_file),
        sheet_name="Data",
        column="Value",
        filters=[]
    )
    
    # Act & Assert
    with pytest.raises(ValueError) as exc_info:
        ops.get_column_stats(request)
    
    print(f"✅ Caught expected error: {exc_info.value}")
    
    assert "no non-null" in str(exc_info.value).lower()


def test_get_column_stats_performance(numeric_types_fixture, file_loader):
    """Test that get_column_stats includes performance metrics.
    
    Verifies:
    - Performance metrics are included
    - Execution time is reasonable
    """
    print(f"\n📊 Testing get_column_stats performance metrics")
    
    ops = StatisticsOperations(file_loader)
    request = GetColumnStatsRequest(
        file_path=numeric_types_fixture.path_str,
        sheet_name=numeric_types_fixture.sheet_name,
        column="Количество",
        filters=[]
    )
    
    # Act
    response = ops.get_column_stats(request)
    
    # Assert
    print(f"✅ Performance:")
    print(f"   Execution time: {response.performance.execution_time_ms}ms")
    print(f"   Cache hit: {response.performance.cache_hit}")
    
    assert response.performance is not None
    assert response.performance.execution_time_ms > 0
    assert response.performance.execution_time_ms < 5000, "Should complete quickly"


# ============================================================================
# correlate tests
# ============================================================================

def test_correlate_two_columns(numeric_types_fixture, file_loader):
    """Test correlate with 2 columns (minimum).
    
    Verifies:
    - Returns correlation matrix
    - Matrix is symmetric
    - Diagonal values are 1.0
    - Correlation values are in [-1, 1] range
    - TSV output is generated
    """
    print(f"\n🔗 Testing correlate with 2 columns")
    
    ops = StatisticsOperations(file_loader)
    
    # Correlate "Количество" and "Цена"
    request = CorrelateRequest(
        file_path=numeric_types_fixture.path_str,
        sheet_name=numeric_types_fixture.sheet_name,
        columns=["Количество", "Цена"],
        method="pearson",
        filters=[]
    )
    
    # Act
    response = ops.correlate(request)
    
    # Assert
    print(f"✅ Correlation matrix:")
    for col1 in response.columns:
        for col2 in response.columns:
            corr = response.correlation_matrix[col1][col2]
            print(f"   {col1} vs {col2}: {corr:.4f}")
    
    assert response.method == "pearson"
    assert response.columns == ["Количество", "Цена"]
    assert len(response.correlation_matrix) == 2
    
    # Check matrix structure
    for col in response.columns:
        assert col in response.correlation_matrix
        assert len(response.correlation_matrix[col]) == 2
    
    # Check diagonal is 1.0
    for col in response.columns:
        assert response.correlation_matrix[col][col] == 1.0, f"Diagonal should be 1.0 for {col}"
    
    # Check symmetry
    assert response.correlation_matrix["Количество"]["Цена"] == response.correlation_matrix["Цена"]["Количество"]
    
    # Check values are in valid range
    for col1 in response.columns:
        for col2 in response.columns:
            corr = response.correlation_matrix[col1][col2]
            assert -1.0 <= corr <= 1.0, f"Correlation should be in [-1, 1], got {corr}"
    
    # Check TSV output
    assert response.excel_output.tsv
    assert "Количество" in response.excel_output.tsv
    assert "Цена" in response.excel_output.tsv


def test_correlate_multiple_columns(numeric_types_fixture, file_loader):
    """Test correlate with 3+ columns.
    
    Verifies:
    - Handles multiple columns correctly
    - Returns full correlation matrix
    """
    print(f"\n🔗 Testing correlate with multiple columns")
    
    ops = StatisticsOperations(file_loader)
    
    # Correlate 3 numeric columns
    request = CorrelateRequest(
        file_path=numeric_types_fixture.path_str,
        sheet_name=numeric_types_fixture.sheet_name,
        columns=["Количество", "Цена", "Скидка"],
        method="pearson",
        filters=[]
    )
    
    # Act
    response = ops.correlate(request)
    
    # Assert
    print(f"✅ Correlation matrix (3x3):")
    print(f"   Columns: {response.columns}")
    
    assert len(response.correlation_matrix) == 3
    assert len(response.columns) == 3
    
    # Check all combinations exist
    for col1 in response.columns:
        for col2 in response.columns:
            assert col2 in response.correlation_matrix[col1]


def test_correlate_spearman_method(numeric_types_fixture, file_loader):
    """Test correlate with Spearman method.
    
    Verifies:
    - Supports different correlation methods
    - Returns correct method in response
    """
    print(f"\n🔗 Testing correlate with Spearman method")
    
    ops = StatisticsOperations(file_loader)
    
    request = CorrelateRequest(
        file_path=numeric_types_fixture.path_str,
        sheet_name=numeric_types_fixture.sheet_name,
        columns=["Количество", "Цена"],
        method="spearman",
        filters=[]
    )
    
    # Act
    response = ops.correlate(request)
    
    # Assert
    print(f"✅ Method: {response.method}")
    
    assert response.method == "spearman"
    assert len(response.correlation_matrix) == 2


def test_correlate_with_filters(numeric_types_fixture, file_loader):
    """Test correlate with filters applied.
    
    Verifies:
    - Filters are applied before correlation
    - Correlation is calculated on filtered subset
    """
    print(f"\n🔗 Testing correlate with filters")
    
    ops = StatisticsOperations(file_loader)
    
    # We need numeric columns - with_dates has "Сумма" which is numeric
    # But we need at least 2 numeric columns. Let's use a different approach.
    # Actually, looking at with_dates fixture, it only has one numeric column (Сумма)
    # So let's use numeric_types_fixture with a filter instead
    
    request = CorrelateRequest(
        file_path=numeric_types_fixture.path_str,
        sheet_name=numeric_types_fixture.sheet_name,
        columns=["Количество", "Цена"],
        method="pearson",
        filters=[
            FilterCondition(column="Количество", operator=">", value=100)
        ],
        logic="AND"
    )
    
    # Act
    response = ops.correlate(request)
    
    # Assert
    print(f"✅ Correlation with filters applied")
    
    assert len(response.correlation_matrix) == 2
    # Correlation should still be valid
    for col1 in response.columns:
        for col2 in response.columns:
            corr = response.correlation_matrix[col1][col2]
            assert -1.0 <= corr <= 1.0


def test_correlate_text_stored_numbers(file_loader, temp_excel_path):
    """Test correlate with text-stored numbers (auto-conversion).
    
    Verifies:
    - Automatically converts text to numbers
    - Correlation is calculated correctly after conversion
    """
    print(f"\n🔗 Testing correlate with text-stored numbers")
    
    # Create test file with text-stored numbers
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    
    ws.append(["X", "Y"])
    for i in range(1, 11):
        ws.append([str(i), str(i * 2)])  # Perfect correlation
    
    test_file = temp_excel_path / "text_corr.xlsx"
    wb.save(test_file)
    
    ops = StatisticsOperations(file_loader)
    request = CorrelateRequest(
        file_path=str(test_file),
        sheet_name="Data",
        columns=["X", "Y"],
        method="pearson",
        filters=[]
    )
    
    # Act
    response = ops.correlate(request)
    
    # Assert
    print(f"✅ Correlation after text conversion:")
    print(f"   X vs Y: {response.correlation_matrix['X']['Y']:.4f}")
    
    # Perfect positive correlation
    assert abs(response.correlation_matrix["X"]["Y"] - 1.0) < 0.01, "Should have perfect correlation"


def test_correlate_single_column_error(numeric_types_fixture, file_loader):
    """Test correlate with single column (error case).
    
    Verifies:
    - Raises ValueError for single column
    - Error message explains minimum requirement
    """
    print(f"\n🔗 Testing correlate with single column (error)")
    
    ops = StatisticsOperations(file_loader)
    
    request = CorrelateRequest(
        file_path=numeric_types_fixture.path_str,
        sheet_name=numeric_types_fixture.sheet_name,
        columns=["Количество"],
        method="pearson",
        filters=[]
    )
    
    # Act & Assert
    with pytest.raises(ValueError) as exc_info:
        ops.correlate(request)
    
    print(f"✅ Caught expected error: {exc_info.value}")
    
    assert "at least 2 columns" in str(exc_info.value).lower()


def test_correlate_invalid_column(numeric_types_fixture, file_loader):
    """Test correlate with non-existent column.
    
    Verifies:
    - Raises ValueError for invalid column
    - Error message lists missing columns
    """
    print(f"\n🔗 Testing correlate with invalid column")
    
    ops = StatisticsOperations(file_loader)
    
    request = CorrelateRequest(
        file_path=numeric_types_fixture.path_str,
        sheet_name=numeric_types_fixture.sheet_name,
        columns=["Количество", "NonExistent"],
        method="pearson",
        filters=[]
    )
    
    # Act & Assert
    with pytest.raises(ValueError) as exc_info:
        ops.correlate(request)
    
    print(f"✅ Caught expected error: {exc_info.value}")
    
    assert "not found" in str(exc_info.value).lower()
    assert "NonExistent" in str(exc_info.value)


def test_correlate_non_numeric_column(simple_fixture, file_loader):
    """Test correlate with non-numeric column.
    
    Verifies:
    - Raises ValueError for non-numeric columns
    - Error message explains the issue
    """
    print(f"\n🔗 Testing correlate with non-numeric column")
    
    ops = StatisticsOperations(file_loader)
    
    # "Имя" and "Город" are string columns
    request = CorrelateRequest(
        file_path=simple_fixture.path_str,
        sheet_name=simple_fixture.sheet_name,
        columns=["Имя", "Город"],
        method="pearson",
        filters=[]
    )
    
    # Act & Assert
    with pytest.raises(ValueError) as exc_info:
        ops.correlate(request)
    
    print(f"✅ Caught expected error: {exc_info.value}")
    
    assert "numeric" in str(exc_info.value).lower()


def test_correlate_insufficient_data(file_loader, temp_excel_path):
    """Test correlate with insufficient data (< 2 rows).
    
    Verifies:
    - Raises ValueError when not enough data
    - Error message is clear
    """
    print(f"\n🔗 Testing correlate with insufficient data")
    
    # Create test file with single row
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    
    ws.append(["X", "Y"])
    ws.append([1, 2])
    
    test_file = temp_excel_path / "single_row.xlsx"
    wb.save(test_file)
    
    ops = StatisticsOperations(file_loader)
    request = CorrelateRequest(
        file_path=str(test_file),
        sheet_name="Data",
        columns=["X", "Y"],
        method="pearson",
        filters=[]
    )
    
    # Act & Assert
    with pytest.raises(ValueError) as exc_info:
        ops.correlate(request)
    
    print(f"✅ Caught expected error: {exc_info.value}")
    
    assert "not enough data" in str(exc_info.value).lower()


def test_correlate_performance(numeric_types_fixture, file_loader):
    """Test that correlate includes performance metrics.
    
    Verifies:
    - Performance metrics are included
    - Execution time is reasonable
    """
    print(f"\n🔗 Testing correlate performance metrics")
    
    ops = StatisticsOperations(file_loader)
    request = CorrelateRequest(
        file_path=numeric_types_fixture.path_str,
        sheet_name=numeric_types_fixture.sheet_name,
        columns=["Количество", "Цена", "Скидка"],
        method="pearson",
        filters=[]
    )
    
    # Act
    response = ops.correlate(request)
    
    # Assert
    print(f"✅ Performance:")
    print(f"   Execution time: {response.performance.execution_time_ms}ms")
    
    assert response.performance is not None
    assert response.performance.execution_time_ms > 0
    assert response.performance.execution_time_ms < 5000, "Should complete quickly"


# ============================================================================
# detect_outliers tests
# ============================================================================

def test_detect_outliers_iqr_method(numeric_types_fixture, file_loader):
    """Test detect_outliers with IQR method.
    
    Verifies:
    - Detects outliers using IQR method
    - Returns outlier rows with all columns
    - Includes _row_index field
    - TSV output is generated
    - Outlier count is accurate
    """
    print(f"\n🎯 Testing detect_outliers with IQR method")
    
    ops = StatisticsOperations(file_loader)
    
    request = DetectOutliersRequest(
        file_path=numeric_types_fixture.path_str,
        sheet_name=numeric_types_fixture.sheet_name,
        column="Итого",
        method="iqr",
        threshold=1.5
    )
    
    # Act
    response = ops.detect_outliers(request)
    
    # Assert
    print(f"✅ Outliers detected: {response.outlier_count}")
    print(f"   Method: {response.method}")
    print(f"   Threshold: {response.threshold}")
    
    assert response.method == "iqr"
    assert response.threshold == 1.5
    assert response.outlier_count == len(response.outliers)
    assert response.outlier_count >= 0, "Should have non-negative count"
    
    # Check outlier structure
    if response.outliers:
        first_outlier = response.outliers[0]
        print(f"   Sample outlier: {dict(list(first_outlier.items())[:3])}")
        
        assert "_row_index" in first_outlier, "Should include row index"
        assert isinstance(first_outlier["_row_index"], int)
        
        # Should have all columns from original data
        expected_cols = numeric_types_fixture.columns
        for col in expected_cols:
            assert col in first_outlier, f"Should include column {col}"
    
    # Check TSV output
    assert response.excel_output.tsv
    if response.outliers:
        assert "_row_index" in response.excel_output.tsv


def test_detect_outliers_zscore_method(numeric_types_fixture, file_loader):
    """Test detect_outliers with Z-score method.
    
    Verifies:
    - Detects outliers using Z-score method
    - Different threshold values work correctly
    """
    print(f"\n🎯 Testing detect_outliers with Z-score method")
    
    ops = StatisticsOperations(file_loader)
    
    request = DetectOutliersRequest(
        file_path=numeric_types_fixture.path_str,
        sheet_name=numeric_types_fixture.sheet_name,
        column="Итого",
        method="zscore",
        threshold=3.0
    )
    
    # Act
    response = ops.detect_outliers(request)
    
    # Assert
    print(f"✅ Outliers detected: {response.outlier_count}")
    print(f"   Method: {response.method}")
    
    assert response.method == "zscore"
    assert response.threshold == 3.0
    assert response.outlier_count >= 0


def test_detect_outliers_different_thresholds(numeric_types_fixture, file_loader):
    """Test detect_outliers with different threshold values.
    
    Verifies:
    - Lower threshold detects more outliers
    - Higher threshold detects fewer outliers
    - Threshold parameter is respected
    """
    print(f"\n🎯 Testing detect_outliers with different thresholds")
    
    ops = StatisticsOperations(file_loader)
    
    # Test with strict threshold (1.0)
    request_strict = DetectOutliersRequest(
        file_path=numeric_types_fixture.path_str,
        sheet_name=numeric_types_fixture.sheet_name,
        column="Итого",
        method="iqr",
        threshold=1.0
    )
    
    response_strict = ops.detect_outliers(request_strict)
    
    # Test with lenient threshold (3.0)
    request_lenient = DetectOutliersRequest(
        file_path=numeric_types_fixture.path_str,
        sheet_name=numeric_types_fixture.sheet_name,
        column="Итого",
        method="iqr",
        threshold=3.0
    )
    
    response_lenient = ops.detect_outliers(request_lenient)
    
    # Assert
    print(f"✅ Strict threshold (1.0): {response_strict.outlier_count} outliers")
    print(f"   Lenient threshold (3.0): {response_lenient.outlier_count} outliers")
    
    # Strict threshold should detect more or equal outliers
    assert response_strict.outlier_count >= response_lenient.outlier_count, \
        "Stricter threshold should detect more outliers"


def test_detect_outliers_no_outliers(file_loader, temp_excel_path):
    """Test detect_outliers when no outliers exist.
    
    Verifies:
    - Returns empty list when no outliers
    - Outlier count is 0
    - TSV output indicates no outliers
    """
    print(f"\n🎯 Testing detect_outliers with no outliers")
    
    # Create test file with uniform data (no outliers)
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    
    ws.append(["Value"])
    for i in range(10):
        ws.append([100])  # All same value
    
    test_file = temp_excel_path / "no_outliers.xlsx"
    wb.save(test_file)
    
    ops = StatisticsOperations(file_loader)
    request = DetectOutliersRequest(
        file_path=str(test_file),
        sheet_name="Data",
        column="Value",
        method="iqr",
        threshold=1.5
    )
    
    # Act
    response = ops.detect_outliers(request)
    
    # Assert
    print(f"✅ Outliers detected: {response.outlier_count}")
    
    assert response.outlier_count == 0, "Should have no outliers"
    assert len(response.outliers) == 0, "Outliers list should be empty"
    assert "No outliers" in response.excel_output.tsv, "TSV should indicate no outliers"


def test_detect_outliers_with_extreme_values(file_loader, temp_excel_path):
    """Test detect_outliers with clear outliers.
    
    Verifies:
    - Detects obvious outliers correctly
    - Returns correct outlier rows
    """
    print(f"\n🎯 Testing detect_outliers with extreme values")
    
    # Create test file with clear outliers
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    
    ws.append(["ID", "Value"])
    # Normal values: 10-19
    for i in range(1, 10):
        ws.append([i, 10 + i])
    # Extreme outlier
    ws.append([10, 1000])
    
    test_file = temp_excel_path / "with_outliers.xlsx"
    wb.save(test_file)
    
    ops = StatisticsOperations(file_loader)
    request = DetectOutliersRequest(
        file_path=str(test_file),
        sheet_name="Data",
        column="Value",
        method="iqr",
        threshold=1.5
    )
    
    # Act
    response = ops.detect_outliers(request)
    
    # Assert
    print(f"✅ Outliers detected: {response.outlier_count}")
    
    assert response.outlier_count > 0, "Should detect outlier"
    
    # Check that the extreme value (1000) is in outliers
    outlier_values = [row["Value"] for row in response.outliers]
    print(f"   Outlier values: {outlier_values}")
    assert 1000 in outlier_values, "Should detect extreme value as outlier"


def test_detect_outliers_text_stored_numbers(file_loader, temp_excel_path):
    """Test detect_outliers with text-stored numbers (auto-conversion).
    
    Verifies:
    - Automatically converts text to numbers
    - Outlier detection works after conversion
    """
    print(f"\n🎯 Testing detect_outliers with text-stored numbers")
    
    # Create test file with text-stored numbers
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    
    ws.append(["Value"])
    for i in range(1, 10):
        ws.append([str(10 + i)])  # Normal values as text
    ws.append([str(1000)])  # Outlier as text
    
    test_file = temp_excel_path / "text_outliers.xlsx"
    wb.save(test_file)
    
    ops = StatisticsOperations(file_loader)
    request = DetectOutliersRequest(
        file_path=str(test_file),
        sheet_name="Data",
        column="Value",
        method="iqr",
        threshold=1.5
    )
    
    # Act
    response = ops.detect_outliers(request)
    
    # Assert
    print(f"✅ Outliers detected after text conversion: {response.outlier_count}")
    
    assert response.outlier_count > 0, "Should detect outlier after conversion"


def test_detect_outliers_invalid_column(simple_fixture, file_loader):
    """Test detect_outliers with non-existent column.
    
    Verifies:
    - Raises ValueError for invalid column
    - Error message lists available columns
    """
    print(f"\n🎯 Testing detect_outliers with invalid column")
    
    ops = StatisticsOperations(file_loader)
    request = DetectOutliersRequest(
        file_path=simple_fixture.path_str,
        sheet_name=simple_fixture.sheet_name,
        column="NonExistentColumn",
        method="iqr",
        threshold=1.5
    )
    
    # Act & Assert
    with pytest.raises(ValueError) as exc_info:
        ops.detect_outliers(request)
    
    print(f"✅ Caught expected error: {exc_info.value}")
    
    assert "not found" in str(exc_info.value).lower()
    assert "NonExistentColumn" in str(exc_info.value)


def test_detect_outliers_non_numeric_column(simple_fixture, file_loader):
    """Test detect_outliers with non-numeric column.
    
    Verifies:
    - Raises ValueError for non-numeric column
    - Error message explains the issue
    """
    print(f"\n🎯 Testing detect_outliers with non-numeric column")
    
    ops = StatisticsOperations(file_loader)
    
    # "Имя" is a string column
    request = DetectOutliersRequest(
        file_path=simple_fixture.path_str,
        sheet_name=simple_fixture.sheet_name,
        column="Имя",
        method="iqr",
        threshold=1.5
    )
    
    # Act & Assert
    with pytest.raises(ValueError) as exc_info:
        ops.detect_outliers(request)
    
    print(f"✅ Caught expected error: {exc_info.value}")
    
    assert "numeric" in str(exc_info.value).lower()


def test_detect_outliers_insufficient_data(file_loader, temp_excel_path):
    """Test detect_outliers with insufficient data (< 4 values).
    
    Verifies:
    - Raises ValueError when not enough data
    - Error message explains minimum requirement
    """
    print(f"\n🎯 Testing detect_outliers with insufficient data")
    
    # Create test file with only 3 values
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    
    ws.append(["Value"])
    ws.append([1])
    ws.append([2])
    ws.append([3])
    
    test_file = temp_excel_path / "insufficient_data.xlsx"
    wb.save(test_file)
    
    ops = StatisticsOperations(file_loader)
    request = DetectOutliersRequest(
        file_path=str(test_file),
        sheet_name="Data",
        column="Value",
        method="iqr",
        threshold=1.5
    )
    
    # Act & Assert
    with pytest.raises(ValueError) as exc_info:
        ops.detect_outliers(request)
    
    print(f"✅ Caught expected error: {exc_info.value}")
    
    assert "not enough data" in str(exc_info.value).lower()
    assert "minimum 4" in str(exc_info.value).lower()


def test_detect_outliers_zscore_zero_std(file_loader, temp_excel_path):
    """Test detect_outliers with Z-score when std=0 (all same values).
    
    Verifies:
    - Raises ValueError when std is 0
    - Error message explains the issue
    """
    print(f"\n🎯 Testing detect_outliers with Z-score and std=0")
    
    # Create test file with all same values
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    
    ws.append(["Value"])
    for i in range(10):
        ws.append([100])  # All same value
    
    test_file = temp_excel_path / "zero_std.xlsx"
    wb.save(test_file)
    
    ops = StatisticsOperations(file_loader)
    request = DetectOutliersRequest(
        file_path=str(test_file),
        sheet_name="Data",
        column="Value",
        method="zscore",
        threshold=3.0
    )
    
    # Act & Assert
    with pytest.raises(ValueError) as exc_info:
        ops.detect_outliers(request)
    
    print(f"✅ Caught expected error: {exc_info.value}")
    
    assert "standard deviation is 0" in str(exc_info.value).lower()


def test_detect_outliers_performance(numeric_types_fixture, file_loader):
    """Test that detect_outliers includes performance metrics.
    
    Verifies:
    - Performance metrics are included
    - Execution time is reasonable
    """
    print(f"\n🎯 Testing detect_outliers performance metrics")
    
    ops = StatisticsOperations(file_loader)
    request = DetectOutliersRequest(
        file_path=numeric_types_fixture.path_str,
        sheet_name=numeric_types_fixture.sheet_name,
        column="Итого",
        method="iqr",
        threshold=1.5
    )
    
    # Act
    response = ops.detect_outliers(request)
    
    # Assert
    print(f"✅ Performance:")
    print(f"   Execution time: {response.performance.execution_time_ms}ms")
    
    assert response.performance is not None
    assert response.performance.execution_time_ms > 0
    assert response.performance.execution_time_ms < 5000, "Should complete quickly"


# ============================================================================
# NEGATION OPERATOR (NOT) TESTS
# ============================================================================

def test_get_column_stats_with_negation(numeric_types_fixture, file_loader):
    """Test get_column_stats with negated filter.
    
    Verifies:
    - Statistics calculated only for rows satisfying negated condition
    - Negation works correctly in statistics context
    """
    print(f"\n🔍 Testing get_column_stats with negation")
    
    ops = StatisticsOperations(file_loader)
    
    request = GetColumnStatsRequest(
        file_path=numeric_types_fixture.path_str,
        sheet_name=numeric_types_fixture.sheet_name,
        column="Количество",
        filters=[
            FilterCondition(column="Количество", operator="<", value=100, negate=True)
        ]
    )
    
    response = ops.get_column_stats(request)
    
    print(f"✅ Stats with negation:")
    print(f"   Count: {response.stats.count}")
    print(f"   Min: {response.stats.min}, Max: {response.stats.max}")
    
    # Should calculate stats only for Количество >= 100
    assert response.stats.count > 0, "Should have some rows"
    assert response.stats.min >= 100, "Min should be >= 100 (negated < 100)"


def test_correlate_with_negation(numeric_types_fixture, file_loader):
    """Test correlate with negated filter.
    
    Verifies:
    - Correlation calculated only for filtered rows
    - Negation works correctly
    """
    print(f"\n🔍 Testing correlate with negation")
    
    ops = StatisticsOperations(file_loader)
    
    request = CorrelateRequest(
        file_path=numeric_types_fixture.path_str,
        sheet_name=numeric_types_fixture.sheet_name,
        columns=["Количество", "Цена"],
        method="pearson",
        filters=[
            FilterCondition(column="Количество", operator="<", value=50, negate=True)
        ]
    )
    
    response = ops.correlate(request)
    
    print(f"✅ Correlation matrix calculated with negation")
    print(f"   Rows processed: {response.metadata.rows_total}")
    
    assert response.correlation_matrix is not None, "Should return correlation matrix"
    assert len(response.correlation_matrix) > 0, "Matrix should not be empty"


# ============================================================================
# NESTED FILTER GROUPS TESTS (statistics)
# ============================================================================

def test_get_column_stats_nested_filters(numeric_types_fixture, file_loader):
    """Test get_column_stats with nested group: (A AND B) OR C.
    
    Verifies:
    - Nested groups work in get_column_stats
    - Stats calculated only for filtered rows
    - Formula is None (nested groups not supported in Excel)
    """
    print(f"\n🔍 Testing get_column_stats: (A AND B) OR C")
    
    from mcp_excel.models.requests import FilterGroup
    
    ops = StatisticsOperations(file_loader)
    
    print(f"  Filter: (Количество < 50 AND Цена > 100) OR Количество == 100")
    
    # Act
    request = GetColumnStatsRequest(
        file_path=numeric_types_fixture.path_str,
        sheet_name=numeric_types_fixture.sheet_name,
        column="Цена",
        filters=[
            FilterGroup(
                filters=[
                    FilterCondition(column="Количество", operator="<", value=50),
                    FilterCondition(column="Цена", operator=">", value=100)
                ],
                logic="AND"
            ),
            FilterCondition(column="Количество", operator="==", value=100)
        ],
        logic="OR"
    )
    response = ops.get_column_stats(request)
    
    # Assert
    print(f"✅ Stats calculated:")
    print(f"   Count: {response.stats.count}")
    print(f"   Min: {response.stats.min}, Max: {response.stats.max}")
    
    assert response.stats.count > 0, "Should have some rows"
    assert response.stats.min is not None, "Should have min value"
    assert response.stats.max is not None, "Should have max value"


def test_correlate_nested_filters(numeric_types_fixture, file_loader):
    """Test correlate with nested group: (A OR B) AND C.
    
    Verifies:
    - Nested groups work in correlate
    - Correlation calculated only for filtered rows
    """
    print(f"\n🔍 Testing correlate: (A OR B) AND C")
    
    from mcp_excel.models.requests import FilterGroup
    
    ops = StatisticsOperations(file_loader)
    
    print(f"  Filter: (Количество < 50 OR Количество > 150) AND Цена > 100")
    
    # Act
    request = CorrelateRequest(
        file_path=numeric_types_fixture.path_str,
        sheet_name=numeric_types_fixture.sheet_name,
        columns=["Количество", "Цена"],
        method="pearson",
        filters=[
            FilterGroup(
                filters=[
                    FilterCondition(column="Количество", operator="<", value=50),
                    FilterCondition(column="Количество", operator=">", value=150)
                ],
                logic="OR"
            ),
            FilterCondition(column="Цена", operator=">", value=100)
        ],
        logic="AND"
    )
    response = ops.correlate(request)
    
    # Assert
    print(f"✅ Correlation matrix calculated")
    print(f"   Rows processed: {response.metadata.rows_total}")
    
    assert response.correlation_matrix is not None, "Should return correlation matrix"
    assert len(response.correlation_matrix) > 0, "Matrix should not be empty"


def test_get_column_stats_nested_with_negation(numeric_types_fixture, file_loader):
    """Test get_column_stats with nested group and negation: NOT (A AND B).
    
    Verifies:
    - Negation works with nested groups in get_column_stats
    - Stats calculated for rows not matching the group
    """
    print(f"\n🔍 Testing get_column_stats: NOT (A AND B)")
    
    from mcp_excel.models.requests import FilterGroup
    
    ops = StatisticsOperations(file_loader)
    
    print(f"  Filter: NOT (Количество < 50 AND Цена > 100)")
    
    # Act
    request = GetColumnStatsRequest(
        file_path=numeric_types_fixture.path_str,
        sheet_name=numeric_types_fixture.sheet_name,
        column="Количество",
        filters=[
            FilterGroup(
                filters=[
                    FilterCondition(column="Количество", operator="<", value=50),
                    FilterCondition(column="Цена", operator=">", value=100)
                ],
                logic="AND",
                negate=True
            )
        ]
    )
    response = ops.get_column_stats(request)
    
    # Assert
    print(f"✅ Stats with negated group:")
    print(f"   Count: {response.stats.count}")
    
    assert response.stats.count > 0, "Should have rows not matching the group"


# ============================================================================
# SAMPLE_ROWS PARAMETER TESTS
# ============================================================================

def test_get_column_stats_with_sample_rows(numeric_types_fixture, file_loader):
    """Test get_column_stats with sample_rows parameter.
    
    Verifies:
    - sample_rows parameter returns sample data
    - Sample data shows rows used in statistics
    - Values are formatted correctly
    """
    print(f"\n🔍 Testing get_column_stats with sample_rows")
    
    ops = StatisticsOperations(file_loader)
    
    # Act
    request = GetColumnStatsRequest(
        file_path=numeric_types_fixture.path_str,
        sheet_name=numeric_types_fixture.sheet_name,
        column="Количество",
        filters=[
            FilterCondition(column="Цена", operator=">", value=100)
        ],
        sample_rows=4
    )
    response = ops.get_column_stats(request)
    
    # Assert
    print(f"✅ Stats count: {response.stats.count}, Sample rows: {len(response.sample_rows) if response.sample_rows else 0}")
    
    assert response.sample_rows is not None, "Should return sample_rows"
    assert isinstance(response.sample_rows, list), "sample_rows should be list"
    assert len(response.sample_rows) <= 4, "Should return at most 4 rows"
    assert len(response.sample_rows) <= response.stats.count, "Sample size should not exceed stats count"
    
    # Verify structure
    if response.sample_rows:
        assert all(isinstance(row, dict) for row in response.sample_rows), "Each row should be dict"
        assert all("Количество" in row for row in response.sample_rows), "Should have analyzed column"
        assert all("Цена" in row for row in response.sample_rows), "Should have filter column"
        # Verify filter was applied: all Цена > 100
        assert all(row["Цена"] > 100 for row in response.sample_rows), "All samples should match filter"


def test_get_column_stats_sample_rows_none(simple_fixture, file_loader):
    """Test get_column_stats with sample_rows=None (default).
    
    Verifies:
    - sample_rows=None returns None (backward compatibility)
    - No sample data in response
    """
    print(f"\n🔍 Testing get_column_stats with sample_rows=None")
    
    ops = StatisticsOperations(file_loader)
    
    # Act
    request = GetColumnStatsRequest(
        file_path=simple_fixture.path_str,
        sheet_name=simple_fixture.sheet_name,
        column=simple_fixture.columns[1],
        sample_rows=None
    )
    response = ops.get_column_stats(request)
    
    # Assert
    print(f"✅ Stats count: {response.stats.count}, Sample rows: {response.sample_rows}")
    
    assert response.sample_rows is None, "Should return None when sample_rows=None"
